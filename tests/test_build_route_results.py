from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


def test_build_xlsm_ships_seeded_route_results() -> None:
    subprocess.run([sys.executable, str(ROOT / "scripts" / "build_xlsm.py")], cwd=ROOT, check=True)

    workbook = openpyxl.load_workbook(XLSM, keep_vba=True)
    try:
        route_results = workbook["route_results"]
        requests = workbook["requests"]

        assert requests.cell(3, 3).value == "SECUI-FW-01;SECUI-FW-02"
        assert route_results.max_row >= 3
        statuses = [route_results.cell(row, 11).value for row in range(2, route_results.max_row + 1)]
        targets = [route_results.cell(row, 10).value for row in range(2, route_results.max_row + 1)]
        assert statuses == ["OK", "OK"]
        assert targets == ["SECUI-FW-01;SECUI-FW-02", "SECUI-FW-01;SECUI-FW-02"]
    finally:
        workbook.close()
