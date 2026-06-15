from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


@pytest.fixture(scope="module")
def xlsm_path() -> Path:
    _ = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "build_xlsm.py")],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    assert XLSM.exists(), "builder did not produce the .xlsm"
    return XLSM


def test_usage_sheet_has_internal_quick_links(xlsm_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    try:
        active = wb.active
        assert active is not None
        active_title = active.title
        usage = wb["usage"]
        targets = {
            "requests",
            "firewalls",
            "firewall_ranges",
            "settings",
            "secui_cli",
            "vendor_cli_templates",
        }

        links: dict[str, str] = {}
        for row in range(1, usage.max_row + 1):
            cell = usage.cell(row=row, column=3)
            if cell.hyperlink is None:
                continue
            target = cell.hyperlink.target
            assert target is not None
            links[str(cell.value)] = target
    finally:
        wb.close()

    assert active_title == "usage"
    assert targets <= set(links)
    for sheet in targets:
        assert links[sheet] == f"#'{sheet}'!A1"

    hidden_targets = {
        "header_aliases",
        "processing_log",
        "service_catalog",
        "sample-request-format",
    }
    assert hidden_targets.isdisjoint(links)
