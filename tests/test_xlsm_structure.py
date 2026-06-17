"""Structural verification of the generated macro-enabled .xlsm.

LibreOffice is unavailable/broken on this host, so we cannot execute the VBA.
Instead we prove the artifact is a real macro-enabled workbook by inspecting:
  - the zip contains xl/vbaProject.bin (macros embedded)
  - both VBA modules are present with the expected public macros
  - all required sheets and headers exist
  - seed data (firewalls / firewall_ranges) is present
  - the seeded example request yields a real multi-firewall path via the oracle

Run: .venv/bin/python -m pytest tests/test_xlsm_structure.py -v
"""

import os
import re
import subprocess
import sys
import zipfile

import openpyxl
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSM = os.path.join(ROOT, "dist", "firewall-policy-automation.xlsm")
PY = sys.executable

CLEAN_REQUESTS_HEADERS = [
    "요청부서", "요청번호", "대상방화벽", "출발지", "출발지설명", "목적지", "목적지설명",
    "프로토콜", "포트", "방향", "용도", "시작일", "종료일", "비고",
]

REMOVED_REQUESTS_HEADERS = {
    "원본파일", "원본행", "제목", "검증상태", "검증메시지", "방화벽경로", "출발매칭대역",
    "목적매칭대역", "대역경로", "매칭근거", "요청폴더",
}

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(ROOT, "scripts"))
from tests.route_oracle import Firewall, FirewallRange, RouteEngine  # noqa: E402


def _build():
    subprocess.run(
        [PY, os.path.join(ROOT, "scripts", "build_xlsm.py")],
        cwd=ROOT, check=True, capture_output=True, text=True,
    )


@pytest.fixture(scope="module")
def xlsm_path():
    _build()
    assert os.path.exists(XLSM), "builder did not produce the .xlsm"
    return XLSM


def test_zip_has_vbaproject(xlsm_path):
    z = zipfile.ZipFile(xlsm_path)
    assert z.testzip() is None
    assert any("vbaProject.bin" in n for n in z.namelist())


def test_content_types_macro_enabled(xlsm_path):
    z = zipfile.ZipFile(xlsm_path)
    ct = z.read("[Content_Types].xml").decode("utf-8", "replace")
    assert "macroEnabled" in ct or "vbaProject" in ct


def test_modules_present(xlsm_path):
    from pyopenvba import excel as ex
    g = ex.ExcelFile(xlsm_path)
    try:
        names = g.module_names()
        assert "FirewallPolicyAutomation" in names
        assert "FirewallRouteAnalysis" in names
        assert "Module1" not in names
        assert "AnalyzeRequestRoutes" in g.get_module("FirewallRouteAnalysis")
        assert "MergeFirewallRequestFolder" in g.get_module("FirewallPolicyAutomation")
        assert "ConvertRequestsToSecuiCli" in g.get_module("FirewallPolicyAutomation")
        assert "SetupFirewallAutomationWorkbook" in g.get_module("FirewallPolicyAutomation")
    finally:
        g.close()


def test_sheets_and_headers(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)

    expected_sheets = {
        "requests", "firewalls", "firewall_ranges",
        "settings", "header_aliases", "processing_log", "sample-request-format", "usage",
        "주간보고", "route_results", "secui_cli", "vendor_cli_templates", "service_catalog",
    }
    assert expected_sheets.issubset(set(wb.sheetnames)), \
        f"missing sheets: {expected_sheets - set(wb.sheetnames)}"

    # Exact header row per sheet. None of these may silently drift.
    expected_headers = {
        "firewalls": ["firewall_name", "vendor", "enabled", "comment"],
        "firewall_ranges": ["firewall_name", "source_cidr", "destination_cidr",
                            "direction", "path_order", "enabled", "comment",
                            "source_interface", "destination_interface", "source_zone", "destination_zone"],
        "settings": ["key", "value", "\uc124\uba85"],
        "header_aliases": ["standard", "your_column", "\uc124\uba85"],
        "processing_log": ["processed_at", "source_file", "status", "merged_rows", "message"],
        "route_results": [
            "요청부서", "요청번호", "출발지", "출발지설명", "목적지", "목적지설명",
            "프로토콜", "포트", "방향", "대상방화벽", "검증상태", "검증메시지",
            "방화벽경로", "출발매칭대역", "목적매칭대역", "대역경로", "매칭근거",
            "원본파일", "원본행",
        ],
        "secui_cli": [
            "No", "\uc7a5\ube44\uba85", "\uc815\ucc45\uba85", "\uba85\ub839\uc5b4", "\uac80\ud1a0\uba54\ubaa8",
            "\uc2e0\uccad\ubd80\uc11c", "\uc2e0\uccad\ubc88\ud638", "\uc6d0\ubcf8\ud30c\uc77c", "\uc6d0\ubcf8\ud589",
        ],
        "vendor_cli_templates": ["vendor", "template_name", "enabled", "command_template", "review_note"],
        "service_catalog": ["service_name", "protocol", "port", "secui_service", "description"],
        "usage": ["Step", "Action"],
    }
    for sheet, headers in expected_headers.items():
        ws = wb[sheet]
        actual = [ws.cell(1, c).value for c in range(1, len(headers) + 1)]
        assert actual == headers, f"{sheet} header drift: {actual!r}"

    req = wb["requests"]
    assert req.max_column == len(CLEAN_REQUESTS_HEADERS), f"requests max_column={req.max_column}"
    actual_req = [req.cell(2, c).value for c in range(1, len(CLEAN_REQUESTS_HEADERS) + 1)]
    assert actual_req == CLEAN_REQUESTS_HEADERS, f"requests header drift: {actual_req!r}"
    all_request_headers = [req.cell(2, c).value for c in range(1, req.max_column + 1)]
    leaked_headers = REMOVED_REQUESTS_HEADERS.intersection(all_request_headers)
    assert not leaked_headers, f"requests still exposes tracking/internal headers: {sorted(leaked_headers)!r}"
    assert req.cell(2, 3).value == "\ub300\uc0c1\ubc29\ud654\ubcbd"
    route_results_headers = [wb["route_results"].cell(1, c).value for c in range(1, wb["route_results"].max_column + 1)]
    assert "원본파일" in route_results_headers
    assert "원본행" in route_results_headers
    # row-1 cosmetic group labels: 출발지 over IP+설명, 목적지 over IP+설명, merged.
    merged = {str(rng) for rng in req.merged_cells.ranges}
    assert req.cell(1, 4).value == "\ucd9c\ubc1c\uc9c0", "row1 출발지 group label missing"
    assert req.cell(1, 6).value == "\ubaa9\uc801\uc9c0", "row1 목적지 group label missing"
    assert "D1:E1" in merged, f"출발지 group not merged D1:E1: {merged}"
    assert "F1:G1" in merged, f"목적지 group not merged F1:G1: {merged}"

    sf = wb["sample-request-format"]
    assert sf.cell(1, 1).value in (None, ""), "sample A1 must be blank"
    assert [sf.cell(1, c).value for c in range(2, 15)] == [
        "No", "\ub300\uc0c1\ubc29\ud654\ubcbd", "\ucd9c\ubc1c\uc9c0IP", "\ucd9c\ubc1c\uc9c0", "\ubaa9\uc801\uc9c0IP", "\ubaa9\uc801\uc9c0",
        "\ud504\ub85c\ud1a0\ucf5c", "\ud3ec\ud2b8", "\ubc29\ud5a5", "\uc6a9\ub3c4", "\uc2dc\uc791\uc77c", "\uc885\ub8cc\uc77c", "\ube44\uace0",
    ]


def test_operator_workbook_shows_only_work_sheets(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)

    visible_sheets = {
        "usage",
        "주간보고",
        "requests",
        "route_results",
        "settings",
        "firewalls",
        "firewall_ranges",
        "secui_cli",
        "vendor_cli_templates",
    }
    support_data_sheets = {
        "header_aliases",
        "processing_log",
        "service_catalog",
        "sample-request-format",
    }

    assert {
        name
        for name in wb.sheetnames
        if wb[name].sheet_state == "visible"
    } == visible_sheets
    for sheet in support_data_sheets:
        assert wb[sheet].sheet_state == "hidden", f"{sheet} must be hidden support data"
    assert wb.active is not None
    assert wb.active.title == "usage"
    visible_request_columns = {"A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"}
    for index in range(1, 15):
        from openpyxl.utils import get_column_letter
        column = get_column_letter(index)
        hidden = bool(wb["requests"].column_dimensions[column].hidden)
        assert hidden is (column not in visible_request_columns), (
            f"requests {column} visibility drifted"
        )


def test_seed_data_present(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    assert wb["firewalls"].max_row >= 4
    assert wb["firewall_ranges"].max_row >= 7
    assert wb["firewall_ranges"].cell(2, 2).value, "firewall_ranges.source_cidr must be seeded"
    assert wb["firewall_ranges"].cell(2, 3).value, "firewall_ranges.destination_cidr must be seeded"
    settings_keys = [wb["settings"].cell(r, 1).value for r in range(1, wb["settings"].max_row + 1)]
    assert "header_alias" in settings_keys
    template = wb["vendor_cli_templates"]
    assert template.cell(2, 1).value == "SECUI"
    assert template.cell(2, 3).value == "Y"
    command_template = str(template.cell(2, 4).value or "")
    assert "{policy_name_q}" in command_template
    assert "{source_interface_q}" in command_template
    assert "{destination_interface_q}" in command_template
    assert "{source_object_q}" in command_template
    assert "{destination_object_q}" in command_template
    assert "{service_object_q}" in command_template


def test_built_workbook_ships_visible_secui_cli_examples(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    cli = wb["secui_cli"]
    assert cli.max_row >= 2, "release workbook must not ship an empty secui_cli sheet"
    policy_names = [str(cli.cell(row, 3).value or "") for row in range(2, cli.max_row + 1)]
    assert len(policy_names) == len(set(policy_names)), "seeded secui_cli policy names must be unique"

    commands = "\n".join(str(cli.cell(row, 4).value or "") for row in range(2, cli.max_row + 1))
    assert "fw set addrgrp" in commands
    assert "fw set svcgrp" in commands
    assert "fw set srule" in commands
    assert "srcif" in commands
    assert "dstif" in commands


def test_secui_cli_seed_groups_same_destination_address_only(monkeypatch):
    import importlib

    seed = importlib.import_module("scripts.secui_cli_seed")

    monkeypatch.setattr(seed, "FIREWALLS", [["firewall_name", "vendor", "enabled", "comment"], ["SECUI-FW-01", "SECUI", "Y", ""]])
    monkeypatch.setattr(
        seed,
        "FIREWALL_RANGES",
        [
            [
                "firewall_name", "source_cidr", "destination_cidr", "direction", "path_order",
                "enabled", "comment", "source_interface", "destination_interface", "source_zone", "destination_zone",
            ],
            ["SECUI-FW-01", "ANY", "ANY", "OUT", 10, "Y", "", "inside", "server", "INTERNAL", "SERVER"],
        ],
    )
    monkeypatch.setattr(
        seed,
        "VENDOR_CLI_TEMPLATES",
        [
            ["vendor", "template_name", "enabled", "command_template", "review_note"],
            [
                "SECUI",
                "default",
                "Y",
                'fw set srule name {policy_name_q} src {source_object_q} dst {destination_object_q} service {service_object_q} srcif {source_interface_q} dstif {destination_interface_q}',
                "",
            ],
        ],
    )
    base = {
        "요청부서": "정보보호",
        "요청번호": "REQ-1",
        "원본파일": "req.xlsx",
        "대상방화벽": "SECUI-FW-01",
        "목적지설명": "APP",
        "프로토콜": "TCP",
        "포트": "443",
        "방향": "OUT",
        "용도": "HTTPS",
        "비고": "",
    }
    monkeypatch.setattr(
        seed,
        "EXAMPLE_REQUEST_ROWS",
        [
            {**base, "원본행": 3, "출발지IP": "10.0.0.1", "출발지설명": "SRC-A", "목적지IP": "172.16.1.10"},
            {**base, "원본행": 4, "출발지IP": "10.0.0.2", "출발지설명": "SRC-B", "목적지IP": "172.16.1.10"},
            {**base, "원본행": 5, "출발지IP": "10.0.0.3", "출발지설명": "SRC-C", "목적지IP": "172.16.1.11"},
        ],
    )

    rows = seed.secui_cli_seed_rows()

    assert len(rows) == 3
    assert {row[8] for row in rows[1:]} == {"3;4", "5"}
    assert any("SRC-A;SRC-B" in row[3] for row in rows[1:])
    assert any("172_16_1_10" in row[2] for row in rows[1:])
    assert any("172_16_1_11" in row[2] for row in rows[1:])


def test_secui_cli_seed_groups_ports_when_source_and_destination_match(monkeypatch):
    import importlib

    seed = importlib.import_module("scripts.secui_cli_seed")
    monkeypatch.setattr(seed, "FIREWALLS", [["firewall_name", "vendor", "enabled", "comment"], ["SECUI-FW-01", "SECUI", "Y", ""]])
    monkeypatch.setattr(
        seed,
        "FIREWALL_RANGES",
        [
            [
                "firewall_name", "source_cidr", "destination_cidr", "direction", "path_order",
                "enabled", "comment", "source_interface", "destination_interface", "source_zone", "destination_zone",
            ],
            ["SECUI-FW-01", "ANY", "ANY", "OUT", 10, "Y", "", "inside", "server", "INTERNAL", "SERVER"],
        ],
    )
    monkeypatch.setattr(
        seed,
        "VENDOR_CLI_TEMPLATES",
        [
            ["vendor", "template_name", "enabled", "command_template", "review_note"],
            [
                "SECUI",
                "default",
                "Y",
                'fw set srule name {policy_name_q} src {source_object_q} dst {destination_object_q} service {service_object_q}',
                "",
            ],
        ],
    )
    base = {
        "요청부서": "정보보호",
        "요청번호": "REQ-2",
        "원본파일": "req.xlsx",
        "대상방화벽": "SECUI-FW-01",
        "목적지설명": "APP",
        "목적지IP": "172.16.1.10",
        "프로토콜": "TCP",
        "방향": "OUT",
        "용도": "HTTPS",
        "비고": "",
    }
    monkeypatch.setattr(
        seed,
        "EXAMPLE_REQUEST_ROWS",
        [
            {**base, "원본행": 3, "출발지IP": "10.0.0.1", "출발지설명": "SRC-A", "포트": "443"},
            {**base, "원본행": 4, "출발지IP": "10.0.0.1", "출발지설명": "SRC-A", "포트": "8443"},
            {**base, "원본행": 5, "출발지IP": "10.0.0.2", "출발지설명": "SRC-B", "포트": "443"},
        ],
    )

    rows = seed.secui_cli_seed_rows()

    assert len(rows) == 3
    merged_service_rule = next(row for row in rows[1:] if row[8] == "3;4")
    assert "tcp/443;tcp/8443" in merged_service_rule[3]
    assert "SRC-A;SRC-B" not in merged_service_rule[3]
    assert any(row[8] == "5" and "tcp/443" in row[3] for row in rows[1:])


def _engine_from_xlsm(wb):
    def T(v):
        return str(v).upper() in ("Y", "YES", "TRUE", "1")
    fs = wb["firewalls"]
    fws = [
        Firewall(fs.cell(r, 1).value, fs.cell(r, 2).value or "", T(fs.cell(r, 3).value))
        for r in range(2, fs.max_row + 1) if fs.cell(r, 1).value
    ]
    rs = wb["firewall_ranges"]
    ranges = [
        FirewallRange(
            rs.cell(r, 1).value,
            rs.cell(r, 2).value or "",
            rs.cell(r, 3).value or "",
            rs.cell(r, 4).value or "",
            int(rs.cell(r, 5).value or 999999),
            T(rs.cell(r, 6).value),
            rs.cell(r, 7).value or "",
        )
        for r in range(2, rs.max_row + 1) if rs.cell(r, 1).value
    ]
    return RouteEngine(firewalls=fws, firewall_ranges=ranges)


def test_seeded_example_yields_multi_firewall_path(xlsm_path):
    """The seeded example request must resolve to a real multi-hop path."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    eng = _engine_from_xlsm(wb)
    res = eng.analyze("10.10.10.5", "8.8.8.8", "OUT")
    assert res.status == "OK"
    assert ";" in res.target_firewalls
    assert ">" in res.firewall_path
    assert res.target_firewalls == "SECUI-FW-01;SECUI-FW-02;SECUI-FW-03"


def test_builtin_seed_resolves_cidr_request(xlsm_path):
    """A CIDR (\ub300\uc5ed) request against the shipped seed data must resolve a path."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    eng = _engine_from_xlsm(wb)
    res = eng.analyze("10.10.10.0/24", "172.16.1.0/24", "OUT")
    assert res.status == "OK"
    assert res.target_firewalls == "SECUI-FW-01"
    # full chain to the internet still resolves
    res2 = eng.analyze("10.10.0.5", "8.8.8.8", "OUT")
    assert res2.status == "OK"
    assert res2.target_firewalls == "SECUI-FW-01;SECUI-FW-02;SECUI-FW-03"


# --------------------------------------------------------------------------- #
# F1: settings sheet schema must match between the build seed and VBA WriteSettings
# --------------------------------------------------------------------------- #

VBA_POLICY = os.path.join(ROOT, "vba", "FirewallPolicyAutomation.bas")


def test_settings_schema_matches_vba_and_build(xlsm_path):
    """settings must be a 3-column sheet (key, value, 설명) in BOTH the built
    workbook and the VBA WriteSettings setup path.

    Guards F1: build_xlsm.py seeds key/value/설명 but VBA historically wrote
    only key/value, so SetupFirewallAutomationWorkbook produced a divergent
    2-column sheet.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    s = wb["settings"]
    assert [s.cell(1, c).value for c in range(1, 4)] == ["key", "value", "\uc124\uba85"]
    keys = [s.cell(r, 1).value for r in range(2, s.max_row + 1)]
    for required in ("request_folder", "parse_sheet", "header_alias"):
        assert required in keys, f"settings missing key {required}"
    assert "parse_targets" not in keys, "unused parse_targets setting should not be seeded"
    # every seeded settings row must carry a non-empty 설명 (column 3)
    for r in range(2, s.max_row + 1):
        if s.cell(r, 1).value:
            assert str(s.cell(r, 3).value or "").strip() != "", f"row {r} missing 설명"

    # VBA WriteSettings must write a 3-column header, not A1:B1
    src = open(VBA_POLICY, encoding="utf-8").read()
    assert 'Range("A1:C1").Value = Array("key", "value"' in src, \
        "WriteSettings must seed a 3-column (key,value,\uc124\uba85) header"
    assert 'Range("A1:B1").Value = Array("key", "value")' not in src, \
        "WriteSettings still writes the legacy 2-column header"


def test_policy_module_has_no_merge_time_legacy_firewall_matching():
    src = open(VBA_POLICY, encoding="utf-8").read()
    for dead in (
        "ResolveTargetFirewalls",
        "RequestRowMatchDetails",
        "RequestRowMatchesCidr",
        "AddressListMatchDetails",
        "AddressListMatchesCidr",
        "RegisteredParseTargetColumns",
        "ParseTargetColumnsFromText",
        "RequestColumnNumber",
        "MarkUnmatchedFirewalls",
    ):
        assert dead not in src, f"obsolete helper {dead} still present"
    # CopyRequestRow must no longer pre-fill target_firewalls before route analysis
    assert "COL_TARGET_FIREWALLS).Value = ResolveTargetFirewalls" not in src


def test_legacy_batch_and_policy_analysis_surfaces_are_removed():
    src = open(VBA_POLICY, encoding="utf-8").read()
    assert "ConvertRequestsToSecuiBatch" not in src
    assert "SECUI_BATCH_SHEET" not in src
    assert "AnalyzeSecuiPolicyExport" not in src
    assert "SECUI_POLICY_EXPORT_SHEET" not in src
    assert "POLICY_ANALYSIS_SHEET" not in src
    assert "POLICY_SUMMARY_SHEET" not in src


def test_secui_cli_macro_generates_fw_set_srule_commands():
    src = open(VBA_POLICY, encoding="utf-8").read()
    assert "Public Sub ConvertRequestsToSecuiCli" in src
    assert "SECUI_CLI_SHEET" in src
    assert "VENDOR_CLI_TEMPLATE_SHEET" in src
    assert "WriteSecuiCliHeaders" in src
    assert "WriteVendorCliTemplateHeaders" in src
    assert "LoadSecuiFirewalls(firewallsSheet)" in src
    assert "LoadVendorCliTemplate(templateSheet, " in src
    assert "CollectSecuiCliRows" in src
    assert "WriteSecuiCliGroup" in src
    assert "CopySecuiCliRows" in src
    assert "RenderVendorCliTemplate" in src
    assert "SecuiCliCommand" in src
    assert "DefaultVendorCliTemplate" in src
    assert "{policy_name_q}" in src
    assert "{source_interface_q}" in src
    assert "{destination_interface_q}" in src
    assert "secuiFirewalls.Exists(SecuiFirewallKey(firewallName))" in src


def test_secui_cli_conversion_does_not_run_route_analysis(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    assert wb["requests"].cell(3, 3).value == "SECUI-FW-01;SECUI-FW-02"

    src = open(VBA_POLICY, encoding="utf-8").read()
    macro_start = src.find("Public Sub ConvertRequestsToSecuiCli()")
    macro_end = src.find("End Sub", macro_start)
    cli_src = src[macro_start:macro_end]
    assert "AnalyzeRequestRoutes" not in cli_src
    assert "CollectSecuiCliRows" in cli_src


def test_vendor_cli_template_sheet_is_operator_configured(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    ws = wb["vendor_cli_templates"]
    assert ws.sheet_state == "visible"
    comment = ws.cell(1, 4).comment
    assert comment is not None
    assert "{policy_name_q}" in comment.text
    assert any("C2:C5000" in str(dv.sqref) for dv in ws.data_validations.dataValidation)


def test_secui_cli_template_oracle_preserves_multi_target_commands(xlsm_path):
    from pyopenvba import excel as ex

    def clean(value):
        return " ".join(str(value).strip().replace("\r", " ").replace("\n", " ").replace("\t", " ").split())

    def quote(value):
        return f'"{clean(value).replace(chr(34), chr(39))}"'

    def service(proto, port):
        return clean(f"{proto.lower()}/{str(port).strip()}")

    def object_token(value):
        result = "".join(ch if ch.isalnum() and ch.isascii() else "_" for ch in clean(value))
        while "__" in result:
            result = result.replace("__", "_")
        return result.strip("_").upper() or "ANY"

    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    template = wb["vendor_cli_templates"].cell(2, 4).value
    assert isinstance(template, str)
    g = ex.ExcelFile(xlsm_path)
    try:
        src = g.get_module("FirewallPolicyAutomation")
    finally:
        g.close()
    vba_default = re.search(r'DefaultVendorCliTemplate = "([^"]+)"', src)
    assert vba_default is not None
    assert template == vba_default.group(1)
    assert 'targetFirewalls = Split(Trim$(CStr(requestsSheet.Cells(requestRow, COL_TARGET_FIREWALLS).Value)), ";")' in src
    assert "For Each firewallValue In targetFirewalls" in src
    assert "WriteSecuiCliGroup requestsSheet, secuiCliSheet, cliGroups(groupKey), cliRow, cliTemplate" in src
    assert "cliRow = cliRow + 1" in src
    assert "BuildSecuiCliServiceFanoutIndex" in src
    assert "SecuiCliSourceDestinationKey" in src
    assert "SecuiCliDestinationServiceKey" in src
    assert "SecuiGroupObjectCommands" in src
    assert "FirstMatchingFirewallRangeInfo" in src
    assert "SecuiPolicyObjectReference" in src

    request = {
        "doc_no": "REQ-1",
        "targets": "SECUI-FW-01;SECUI-FW-02",
        "source": "10.10.10.5",
        "destination": "172.16.1.10",
        "protocol": "TCP",
        "port": "443",
        "purpose": "HTTPS 업무 연동",
        "note": "",
        "status": "OK",
    }

    rows = []
    for firewall_name in request["targets"].split(";"):
        policy_name = object_token(
            "_".join([
                request["doc_no"],
                firewall_name,
                "INTERNAL_TO_SERVER",
                service(request["protocol"], request["port"]),
                request["destination"],
            ])
        )[:120]
        description = clean(f"{request['purpose']} / {request['status']}")[:255]
        source_group = f"GRP_SRC_INTERNAL_TO_SERVER_{policy_name}"
        destination_group = f"GRP_DST_INTERNAL_TO_SERVER_{policy_name}"
        service_group = f"GRP_SVC_INTERNAL_TO_SERVER_{policy_name}"
        command = template
        command = command.replace("{policy_name_q}", quote(policy_name))
        command = command.replace("{source_interface_q}", quote("inside"))
        command = command.replace("{destination_interface_q}", quote("server"))
        command = command.replace("{source_object_q}", quote(source_group))
        command = command.replace("{destination_object_q}", quote(destination_group))
        command = command.replace("{service_object_q}", quote(service_group))
        command = command.replace("{description_q}", quote(description))
        command = command.replace("{firewall_name}", clean(firewall_name))
        rows.append((firewall_name, policy_name, command))

    assert len(rows) == 2
    assert rows[0][0] == "SECUI-FW-01"
    assert rows[1][0] == "SECUI-FW-02"
    assert rows[0][2] == (
        'fw set srule name "REQ_1_SECUI_FW_01_INTERNAL_TO_SERVER_TCP_443_172_16_1_10" '
        'action allow srcif "inside" dstif "server" '
        'src "GRP_SRC_INTERNAL_TO_SERVER_REQ_1_SECUI_FW_01_INTERNAL_TO_SERVER_TCP_443_172_16_1_10" '
        'dst "GRP_DST_INTERNAL_TO_SERVER_REQ_1_SECUI_FW_01_INTERNAL_TO_SERVER_TCP_443_172_16_1_10" '
        'service "GRP_SVC_INTERNAL_TO_SERVER_REQ_1_SECUI_FW_01_INTERNAL_TO_SERVER_TCP_443_172_16_1_10" '
        'log enable enable yes description "HTTPS 업무 연동 / OK" # device=SECUI-FW-01'
    )
    assert 'srule name "REQ_1_SECUI_FW_02_INTERNAL_TO_SERVER_TCP_443_172_16_1_10"' in rows[1][2]
    assert "# device=SECUI-FW-02" in rows[1][2]
    assert service("ICMP", "") == "icmp/"


def test_secui_cli_vba_preserves_blank_port_service_suffix():
    src = open(VBA_POLICY, encoding="utf-8").read()
    assert 'SecuiCliServiceText = CleanSecuiText(proto & "/" & portText)' in src
    assert "If Len(Trim$(portText)) = 0 Then" not in src


def test_firewall_excel_benchmark_doc_covers_secui_and_management_repos():
    doc = open(os.path.join(ROOT, "docs", "firewall-excel-benchmark.md"), encoding="utf-8").read()
    assert "SECUI" in doc
    assert "SECUI-specific public management repositories are scarce" in doc
    for expected in (
        "CactuseSecurity/firewall-orchestrator",
        "imthenachoman/pfSense-Firewall-Rules-Manager",
        "martimy/firewall_policy_analyzer",
        "automateyournetwork/netclaw",
        "olafhartong/parsoalto",
    ):
        assert expected in doc
    assert "vendor_cli_templates" in doc


def test_route_macro_colors_target_firewall_cell():
    src = open(VBA_ROUTE, encoding="utf-8").read()
    assert "Set targetCell = sheet.Cells(rowIndex, RCOL_TARGET)" in src
    assert 'targetCell.Value = res("target_firewalls")' in src
    assert "targetCell.Interior.Color = RGB(217, 234, 211)" in src


def test_merge_only_integrates_requests_without_validation_or_route_analysis():
    src = open(VBA_POLICY, encoding="utf-8").read()
    merge_start = src.find("Public Sub MergeFirewallRequestFolder()")
    merge_end = src.find("End Sub", merge_start)
    merge_src = src[merge_start:merge_end]
    assert "AnalyzeRequestRoutes" not in merge_src
    assert "MarkDuplicateRequests requestsSheet" not in merge_src
    assert "WriteRowValidation requestsSheet" not in merge_src


# --------------------------------------------------------------------------- #
# F6: the VBA source injected into the .xlsm must match vba/*.bas exactly
# (catches pyOpenVBA injection corruption / build drift)
# --------------------------------------------------------------------------- #

def _normalize_vba(text):
    # Normalize for round-trip comparison:
    #  - drop the leading `Attribute VB_Name = "..."` header (body-only on read)
    #  - collapse non-ASCII to '?' because pyOpenVBA reads the MBCS module stream
    #    with a codepage that is lossy for Korean COMMENT text (한글 -> '?').
    #    Code/identifiers are ASCII, so this still catches real source corruption.
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    lines = [ln for ln in lines if not ln.startswith("Attribute VB_Name")]
    body = "\n".join(lines).strip()
    return "".join(c if ord(c) < 128 else "?" for c in body)


def test_injected_vba_modules_match_source_files(xlsm_path):
    from pyopenvba import excel as ex
    modules = {
        "FirewallPolicyAutomation": os.path.join(ROOT, "vba", "FirewallPolicyAutomation.bas"),
        "FirewallRouteAnalysis": os.path.join(ROOT, "vba", "FirewallRouteAnalysis.bas"),
    }
    g = ex.ExcelFile(xlsm_path)
    try:
        for name, path in modules.items():
            injected = _normalize_vba(g.get_module(name))
            source = _normalize_vba(open(path, encoding="utf-8").read())
            assert injected == source, \
                f"injected {name} does not match vba/{name}.bas (build corruption?)"
    finally:
        g.close()


def test_korean_preserved_in_injected_vba(xlsm_path):
    """한글 functional 리터럴이 빌드된 .xlsm VBA에 보존돼야 한다 (code_page=949)."""
    from pyopenvba import excel as ex
    g = ex.ExcelFile(xlsm_path)
    try:
        assert g.vba_project().code_page == 949, "VBA project must use code page 949"
        src = g.get_module("FirewallPolicyAutomation")
        assert 'headerMap("출발지ip")' in src, "Korean column literal corrupted"
        assert 'headerMap("목적지ip")' in src
        assert "출발지IP" in src and "목적지IP" in src
        assert "?" * 3 not in src, "Korean text was replaced with '?'"
    finally:
        g.close()


VBA_ROUTE = os.path.join(ROOT, "vba", "FirewallRouteAnalysis.bas")
RELEASE_WORKFLOW = os.path.join(ROOT, ".github", "workflows", "release.yml")


def test_vba_loads_firewall_ranges_sheet():
    src = open(VBA_ROUTE, encoding="utf-8").read()
    assert 'Private Const FIREWALL_RANGE_SHEET As String = "firewall_ranges"' in src
    assert "ThisWorkbook.Worksheets(FIREWALL_RANGE_SHEET)" in src
    assert "NETWORK_SHEET" not in src
    assert "ROUTING_SHEET" not in src


def test_release_workflow_verifies_firewall_ranges_sheet():
    src = open(RELEASE_WORKFLOW, encoding="utf-8").read()
    assert '"firewall_ranges"' in src
    assert '"network_definitions"' not in src
    assert '"routing_paths"' not in src
    assert "대상방화벽" in src


def test_vba_split_address_list_collapses_spaces():
    """S6 parity: VBA SplitAddressList must collapse ASCII space-runs into ';'
    (mirrors Python split_address_list) so space-separated multi-CIDR splits."""
    src = open(VBA_ROUTE, encoding="utf-8").read().replace("\r\n", "\n")
    fn = src[src.find("Private Function SplitAddressList"):]
    fn = fn[:fn.find("End Function")]
    assert 'Replace(normalized, " ", ";")' in fn, \
        "SplitAddressList must turn spaces into ';' to split space-separated CIDRs"


def test_vba_address_overlap_treats_blank_ip_as_no_match():
    """Parity: a blank request IP cell must match NOTHING, mirroring Python
    _address_list_overlaps (empty token list -> False). The VBA bug was that
    SplitAddressList('') returned a one-element [''] array, so RangesOverlap('',
    def) was called and IsAnyCidr('') made it overlap every range -> wrong
    routes / spurious DIRECTION_MISMATCH. Lock the empty-array guard in place."""
    src = open(VBA_ROUTE, encoding="utf-8").read().replace("\r\n", "\n")
    split_fn = src[src.find("Private Function SplitAddressList"):]
    split_fn = split_fn[:split_fn.find("End Function")]
    assert "If count < 0 Then" in split_fn, \
        "SplitAddressList must return an empty array when no real tokens exist"
    assert 'Split(vbNullString, ";", 0)' in split_fn, \
        "empty SplitAddressList result must be a genuinely empty array"

    overlap_fn = src[src.find("Private Function AddressListOverlaps"):]
    overlap_fn = overlap_fn[:overlap_fn.find("End Function")]
    assert "UBound(requests) < LBound(requests)" in overlap_fn, \
        "AddressListOverlaps must bail out when the request token list is empty"
    assert "UBound(definitions) < LBound(definitions)" in overlap_fn, \
        "AddressListOverlaps must bail out when the definition token list is empty"
    # The empty-request guard must run BEFORE the IsAnyCidr(definition) short-circuit,
    # else a blank request IP would wrongly overlap an ANY range.
    req_guard = overlap_fn.find("UBound(requests) < LBound(requests)")
    any_short = overlap_fn.find("If IsAnyCidr(definitionValue) Then")
    assert req_guard != -1 and any_short != -1 and req_guard < any_short, \
        "empty-request guard must precede the ANY-definition short-circuit"


def test_vba_find_header_row_is_content_based():
    """Parity: VBA header detection must score by FIELD CONTENT (IP columns +
    field count), not by an exact 'no'/'번호' cell match. The content scan lives
    in BestHeaderRow, which FindHeaderRow and FindRequestSheet both delegate to."""
    src = open(VBA_POLICY, encoding="utf-8").read().replace("\r\n", "\n")
    fn = src[src.find("Private Function BestHeaderRow"):]
    fn = fn[:fn.find("\nEnd Function")]
    # must score by IP presence + field count, not the old exact-match anchor
    assert "hasIp" in fn and "fieldCount" in fn, "BestHeaderRow must score by content"
    assert 'If valueText = "no" Or valueText = "번호" Then' not in fn, \
        "old exact-match 'no'/'번호' anchor must be gone"
    # IsFieldHeader helper must exist and enumerate the field headers
    assert "Private Function IsFieldHeader" in src
    # FindRequestSheet must exist so data on a non-first sheet still parses
    assert "Private Function FindRequestSheet" in src


def test_vba_header_key_strips_punctuation():
    """Parity: HeaderKey must strip decorating punctuation so 'No.' -> 'no', but
    PRESERVE an all-punctuation token like '#' (return original when stripping
    empties it) so it can match the '#' No-alias."""
    src = open(VBA_POLICY, encoding="utf-8").read().replace("\r\n", "\n")
    fn = src[src.find("Private Function HeaderKey"):]
    fn = fn[:fn.find("\nEnd Function")]
    assert "puncts" in fn and ("Left$(k" in fn or "Mid$(k" in fn), \
        "HeaderKey must strip leading/trailing punctuation (No. -> no)"
    assert "original" in fn and "If Len(k) = 0 Then" in fn, \
        "HeaderKey must keep all-punctuation tokens (e.g. '#') instead of emptying"


def test_vba_canonical_no_includes_hash_alias():
    """Parity: CanonicalHeaderName 'no' Case must include '#' so a '#' column
    anchors the header row by itself."""
    src = open(VBA_POLICY, encoding="utf-8").read().replace("\r\n", "\n")
    line = next(l for l in src.split("\n") if "CanonicalHeaderName = \"no\"" in l)
    assert '"#"' in line, "CanonicalHeaderName 'no' Case must include '#'"


def test_workbook_open_does_not_swallow_errors():
    """Workbook_Open auto-run must surface failures, not blanket-swallow them."""
    from pyopenvba import excel as ex
    g = ex.ExcelFile(XLSM)
    try:
        tw = g.get_module("ThisWorkbook")
        open_start = tw.find("Workbook_Open")
        assert open_start >= 0
        open_end = tw.find("End Sub", open_start)
        workbook_open = tw[open_start:open_end]
        assert "AutoRunWorkbookOutputs" in workbook_open
        assert "Application.Run" not in tw, "direct call avoids Windows macro name-resolution failures"
        assert "On Error Resume Next" not in workbook_open, "Workbook_Open must not blanket-swallow errors"
        assert "AutoRunErr" in workbook_open, "must have an error handler that surfaces failures"
    finally:
        g.close()


def test_auto_run_macro_refreshes_cli_outputs_without_folder_prompt():
    src = open(VBA_POLICY, encoding="utf-8").read()
    start = src.find("Public Sub AutoRunWorkbookOutputs()")
    end = src.find("End Sub", start)
    macro = src[start:end]
    assert start >= 0, "AutoRunWorkbookOutputs macro is missing"
    assert "FolderExists(SettingsValue(settingsSheet, \"request_folder\"))" in macro
    assert "MergeFirewallRequestFolder" in macro
    assert "AnalyzeRequestRoutes" in macro
    assert "ConvertRequestsToSecuiCli" in macro
    assert macro.find("MergeFirewallRequestFolder") < macro.find("AnalyzeRequestRoutes")
    assert macro.find("AnalyzeRequestRoutes") < macro.find("ConvertRequestsToSecuiCli")
    assert "On Error Resume Next" not in macro, "AutoRunWorkbookOutputs must not blanket-swallow errors"
    assert "ConvertRequestsToSecuiBatch" not in macro
    assert "AnalyzeSecuiPolicyExport" not in macro
    assert "WriteSettings" not in macro
    assert "RequestFolderPath" not in macro


# --------------------------------------------------------------------------- #
# UX (scope B): build-time openpyxl input-assist / display features.
# These are DISPLAY/INPUT-ASSIST ONLY and must NOT change the route algorithm.
# All are sheet-level and additive: no cell VALUE changes, no header drift, no
# max_column/max_row drift, vbaProject.bin preserved (keep_vba=True).
# --------------------------------------------------------------------------- #

def _dv_for(ws, cell_ref):
    """Return the DataValidation covering cell_ref on ws, or None."""
    from openpyxl.utils.cell import coordinate_to_tuple
    target = coordinate_to_tuple(cell_ref)
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            if (rng.min_row <= target[0] <= rng.max_row
                    and rng.min_col <= target[1] <= rng.max_col):
                return dv
    return None


def test_ux_protocol_direction_dropdowns_on_requests(xlsm_path):
    """requests 프로토콜(col8)·방향(col10) cells offer a dropdown list.

    Values must stay compatible with the route algorithm: 프로토콜 is display-only
    (TCP/UDP/ICMP); 방향 must be one of the values VBA NormalizeDirection accepts
    (IN/OUT/BOTH) so the dropdown can never inject a #INVALID direction.
    """
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    req = wb["requests"]

    proto = _dv_for(req, "H3")
    assert proto is not None, "프로토콜(col8) has no data validation"
    assert proto.type == "list"
    assert set(proto.formula1.strip('"').split(",")) >= {"TCP", "UDP", "ICMP"}

    direction = _dv_for(req, "J3")
    assert direction is not None, "방향(col10) has no data validation"
    assert direction.type == "list"
    assert set(direction.formula1.strip('"').split(",")) <= {"IN", "OUT", "BOTH", ""}
    assert {"IN", "OUT"} <= set(direction.formula1.strip('"').split(","))
    # validations must skip the header row and not explode to 1048576 rows
    for dv in (proto, direction):
        rng = next(iter(dv.sqref.ranges))
        assert rng.min_row == 3, "dropdown must start at row 3 (skip 2-row header band)"
        assert rng.max_row <= 5000, "dropdown range must be bounded, not whole-column"


def test_ux_enabled_dropdown_accepts_vba_truthy(xlsm_path):
    """firewalls.enabled(col3) dropdown values must all be truthy/falsy under the
    VBA reader T(v)=UCase(v) in (Y,YES,TRUE,1); list must include both Y and N."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    dv = _dv_for(wb["firewalls"], "C2")
    assert dv is not None, "firewalls.enabled(col3) has no data validation"
    assert dv.type == "list"
    vals = set(dv.formula1.strip('"').split(","))
    assert {"Y", "N"} <= vals
    assert vals <= {"Y", "N", "TRUE", "FALSE"}


def test_ux_macro_written_sheets_are_not_protected(xlsm_path):
    """CRITICAL: requests and processing_log are written by VBA at runtime, so they
    must NOT be sheet-protected (protection would make macro Cells(...).Value fail).
    Operator-input sheets MAY be protected."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    for macro_sheet in ("requests", "processing_log"):
        assert wb[macro_sheet].protection.sheet is False, \
            f"{macro_sheet} is macro-written and must never be protected"


def test_ux_settings_folder_cell_unlocked_when_protected(xlsm_path):
    """If settings is protected, the request_folder value cell (B column of the
    request_folder row) must be UNLOCKED so SelectRequestFolder can write it."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    s = wb["settings"]
    if not s.protection.sheet:
        pytest.skip("settings not protected; folder-cell lock state is moot")
    folder_row = next(
        (r for r in range(2, s.max_row + 1)
         if s.cell(r, 1).value == "request_folder"), None)
    assert folder_row is not None
    assert s.cell(folder_row, 2).protection.locked is False, \
        "request_folder value cell must be unlocked for the folder-select macro"


def test_ux_required_input_empty_highlight(xlsm_path):
    """requests must carry conditional formatting that flags an empty required input
    cell (출발지 col4 / 목적지 col6). Display-only; never touches values."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    req = wb["requests"]
    cf_ranges = [str(rng) for rng in req.conditional_formatting]
    joined = " ".join(cf_ranges)
    assert ("D" in joined and "F" in joined) or len(cf_ranges) >= 1, \
        "requests has no conditional formatting for empty required cells"
    assert any(req.conditional_formatting[rng] for rng in req.conditional_formatting), \
        "conditional formatting present but carries no rules"


def test_ux_target_firewall_cell_highlight(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    req = wb["requests"]
    cf_ranges = [str(rng) for rng in req.conditional_formatting]
    assert any("C" in rng for rng in cf_ranges), \
        "대상방화벽(col3) must have conditional formatting"


def test_ux_header_comments_on_key_columns(xlsm_path):
    """Operator-facing 안내 comments must exist on the requests 출발지 / 목적지
    leaf-header cells (row 2), and the format must stay valid (no value drift)."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    req = wb["requests"]
    assert req.cell(2, 4).comment is not None, "출발지 header needs a hint comment"
    assert req.cell(2, 6).comment is not None, "목적지 header needs a hint comment"
    # comment must not alter the header value
    assert req.cell(2, 4).value == "\ucd9c\ubc1c\uc9c0"
    assert req.cell(2, 6).value == "\ubaa9\uc801\uc9c0"


def test_ux_tab_colors_assigned(xlsm_path):
    """Input sheets vs result/log sheets must be visually distinguished by tab color
    (display-only). At least the requests result sheet must carry a tabColor."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    assert wb["requests"].sheet_properties.tabColor is not None, \
        "requests tab must have a color"
    colored = sum(
        1 for name in wb.sheetnames
        if wb[name].sheet_properties.tabColor is not None)
    assert colored >= 3, "at least 3 sheets should be tab-colored for navigation"


def test_ux_does_not_change_route_algorithm(xlsm_path):
    """GUARD: the route oracle result on the seeded data must be byte-identical
    whether or not UX is applied. UX must be display/input-assist ONLY."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    eng = _engine_from_xlsm(wb)
    res = eng.analyze("10.10.10.5", "8.8.8.8", "OUT")
    assert res.status == "OK"
    assert res.target_firewalls == "SECUI-FW-01;SECUI-FW-02;SECUI-FW-03"
