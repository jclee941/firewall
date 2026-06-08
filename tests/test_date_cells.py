"""실제 엑셀 날짜 셀 처리 검증 (시작일/종료일).

신청서의 시작일/종료일이 문자열이 아니라 **실제 엑셀 날짜 타입**일 때,
오라클과 VBA가 동일하게 yyyy-mm-dd 로 저장해야 한다 (로케일 비의존).
- 문자열 날짜는 그대로 유지
- 다른 컬럼(IP/포트/프로토콜 등)은 영향 없음
- 시작일/종료일이 date 타입일 때만 yyyy-mm-dd 포맷

Run: .venv/bin/python -m pytest tests/test_date_cells.py -v
"""

import datetime
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(__file__))

from request_parser_oracle import parse_request_sheet, sheet_to_filled_rows  # noqa: E402


def _build(tmp_path, name, build_fn):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_fn(ws)
    p = tmp_path / name
    wb.save(p)
    return p


_HDR = ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜",
        "포트", "방향", "용도", "시작일", "종료일", "비고"]


def test_real_date_cells_formatted_yyyymmdd(tmp_path):
    def build(ws):
        for i, h in enumerate(_HDR, 1):
            ws.cell(1, i, h)
        ws.cell(2, 1, 1); ws.cell(2, 2, "10.10.10.5"); ws.cell(2, 3, "PC")
        ws.cell(2, 4, "10.20.20.5"); ws.cell(2, 5, "DMZ"); ws.cell(2, 6, "tcp")
        ws.cell(2, 7, 443)  # int port
        ws.cell(2, 8, "OUT"); ws.cell(2, 9, "업무")
        ws.cell(2, 10, datetime.datetime(2026, 1, 1))    # real date cell
        ws.cell(2, 11, datetime.date(2026, 12, 31))      # real date (no time)
        ws.cell(2, 12, "")

    p = _build(tmp_path, "dates.xlsx", build)
    parsed = parse_request_sheet(sheet_to_filled_rows(openpyxl.load_workbook(p).active))
    assert len(parsed) == 1
    r = parsed[0]
    # date cells -> yyyy-mm-dd, no time component
    assert r["start_date"] == "2026-01-01"
    assert r["end_date"] == "2026-12-31"
    # other columns unaffected
    assert r["port"] == "443"
    assert r["protocol"] == "TCP"
    assert r["source_ip"] == "10.10.10.5"


def test_string_dates_unchanged(tmp_path):
    def build(ws):
        for i, h in enumerate(_HDR, 1):
            ws.cell(1, i, h)
        ws.cell(2, 1, 1); ws.cell(2, 2, "10.10.10.5"); ws.cell(2, 3, "PC")
        ws.cell(2, 4, "10.20.20.5"); ws.cell(2, 5, "DMZ"); ws.cell(2, 6, "TCP")
        ws.cell(2, 7, "443"); ws.cell(2, 8, "OUT"); ws.cell(2, 9, "업무")
        ws.cell(2, 10, "2026-01-01")  # string date stays as-is
        ws.cell(2, 11, "2026.12.31")  # non-ISO string also untouched
        ws.cell(2, 12, "")

    p = _build(tmp_path, "strdates.xlsx", build)
    parsed = parse_request_sheet(sheet_to_filled_rows(openpyxl.load_workbook(p).active))
    r = parsed[0]
    assert r["start_date"] == "2026-01-01"
    assert r["end_date"] == "2026.12.31"  # not reformatted
