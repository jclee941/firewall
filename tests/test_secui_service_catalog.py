from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSM = os.path.join(ROOT, "dist", "firewall-policy-automation.xlsm")
PY = sys.executable


def _build() -> str:
    subprocess.run(
        [PY, os.path.join(ROOT, "scripts", "build_xlsm.py")],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return XLSM


def test_secui_service_catalog_sheet_is_seeded():
    xlsm = _build()
    wb = openpyxl.load_workbook(xlsm, keep_vba=True)

    assert "service_catalog" in wb.sheetnames
    ws = wb["service_catalog"]
    headers = [ws.cell(1, col).value for col in range(1, 6)]
    assert headers == [
        "service_name",
        "protocol",
        "port",
        "secui_service",
        "description",
    ]
    rows = {
        ws.cell(row, 1).value: [ws.cell(row, col).value for col in range(2, 5)]
        for row in range(2, ws.max_row + 1)
    }
    assert rows["HTTPS"] == ["TCP", "443", "tcp/443"]
    assert rows["DNS-UDP"] == ["UDP", "53", "udp/53"]


def test_secui_protocol_port_headers_point_to_catalog_without_restricting_input():
    xlsm = _build()
    wb = openpyxl.load_workbook(xlsm, keep_vba=True)
    req = wb["requests"]

    protocol_comment = req.cell(2, 12).comment
    port_comment = req.cell(2, 13).comment
    assert protocol_comment is not None
    assert port_comment is not None
    joined = f"{protocol_comment.text}\n{port_comment.text}"
    assert "SECUI" in joined
    assert "service_catalog" in joined
    assert "tcp/443" in joined
    assert "udp/53" in joined
    assert all("M3" not in validation.cells for validation in req.data_validations.dataValidation)


def test_vba_setup_seeds_secui_service_catalog():
    src = Path(ROOT, "vba", "FirewallPolicyAutomation.bas").read_text(encoding="utf-8")

    assert 'Private Const SERVICE_CATALOG_SHEET As String = "service_catalog"' in src
    assert "WriteServiceCatalogHeaders serviceCatalogSheet" in src
    assert 'Array("HTTPS", "TCP", "443", "tcp/443", "웹 HTTPS")' in src
