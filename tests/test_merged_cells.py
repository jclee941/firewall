"""병합 셀(merged cells) 처리 검증.

신청서 양식은 흔히 병합 셀을 씁니다:
  - 헤더 위 제목/그룹 헤더 병합 (이미 find_header_row가 우회)
  - 데이터 셀 세로/가로 병합 (한 신청건이 여러 행/열에 걸침)

규칙(Excel/VBA MergeArea 동작 미러):
  병합 영역의 좌상단 셀 값을 영역 전체로 전파(fill)한 뒤 파싱한다.
  이래야 세로 병합된 출발지IP 등이 모든 행에서 읽혀 누락이 없다.

Run: .venv/bin/python -m pytest tests/test_merged_cells.py -v
"""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(__file__))

from request_parser_oracle import parse_request_sheet  # noqa: E402
from route_oracle import Firewall, Network, RouteEngine, RoutingPath  # noqa: E402


@pytest.fixture
def engine():
    nets = [
        Network("업무PC망", "10.10.0.0/16", "internal"),
        Network("서버망", "172.16.1.0/24", "server"),
        Network("중간망", "10.30.0.0/16", "transit"),
        Network("DMZ망", "10.20.0.0/16", "dmz"),
        Network("외부", "0.0.0.0/0", "outside"),
    ]
    fws = [Firewall("SECUI-FW-01"), Firewall("SECUI-FW-02"), Firewall("SECUI-FW-03")]
    rps = [
        RoutingPath("SECUI-FW-01", "internal", "server", "eth1", "eth2", 10),
        RoutingPath("SECUI-FW-01", "internal", "transit", "eth1", "eth3", 20),
        RoutingPath("SECUI-FW-02", "transit", "dmz", "eth1", "eth2", 30),
        RoutingPath("SECUI-FW-03", "dmz", "outside", "eth1", "eth2", 40),
    ]
    return RouteEngine(networks=nets, firewalls=fws, routing_paths=rps)


def _rows_from(path):
    """Read sheet into list[list], filling merged-cell values across the range.

    This is the production helper under test: it must propagate the top-left
    value of each merged range to every cell in that range BEFORE parsing,
    mirroring how Excel/VBA MergeArea exposes the value.
    """
    from request_parser_oracle import sheet_to_filled_rows  # under test
    wb = openpyxl.load_workbook(path)
    return sheet_to_filled_rows(wb.active)


def _build(tmp_path, name, build_fn):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_fn(ws)
    p = tmp_path / name
    wb.save(p)
    return p


# --------------------------------------------------------------------------- #
# S2: 데이터 세로 병합 — 출발지IP가 B2:B3 병합. 둘째 행도 같은 출발지IP로 채워져야 함.
# --------------------------------------------------------------------------- #

def test_vertical_merge_fills_down(tmp_path, engine):
    def build(ws):
        hdr = ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜",
               "포트", "방향", "용도", "시작일", "종료일", "비고"]
        for i, h in enumerate(hdr, 1):
            ws.cell(1, i, h)
        # row2~3: 출발지IP / 출발지 세로 병합, 목적지는 행마다 다름
        ws.cell(2, 1, 1); ws.cell(2, 2, "10.10.10.5"); ws.merge_cells("B2:B3")
        ws.cell(2, 3, "PC"); ws.merge_cells("C2:C3")
        ws.cell(2, 4, "10.20.20.5"); ws.cell(2, 6, "TCP"); ws.cell(2, 7, "443"); ws.cell(2, 8, "OUT"); ws.cell(2, 9, "a")
        ws.cell(3, 4, "172.16.1.10"); ws.cell(3, 6, "TCP"); ws.cell(3, 7, "22"); ws.cell(3, 8, "OUT"); ws.cell(3, 9, "b")

    p = _build(tmp_path, "vmerge.xlsx", build)
    parsed = parse_request_sheet(_rows_from(p))
    assert len(parsed) == 2
    # both rows must carry the merged 출발지IP
    assert parsed[0]["source_ip"] == "10.10.10.5"
    assert parsed[1]["source_ip"] == "10.10.10.5"   # filled from merge
    assert parsed[1]["dest_ip"] == "172.16.1.10"
    # and both route correctly
    r0 = engine.analyze(parsed[0]["source_ip"], parsed[0]["dest_ip"], parsed[0]["direction"])
    r1 = engine.analyze(parsed[1]["source_ip"], parsed[1]["dest_ip"], parsed[1]["direction"])
    assert r0.status == "OK" and r0.target_firewalls == "SECUI-FW-01;SECUI-FW-02"
    assert r1.status == "OK" and r1.target_firewalls == "SECUI-FW-01"


# --------------------------------------------------------------------------- #
# S1: 헤더 위 병합 제목/그룹 헤더 — 진짜 헤더 행을 찾고 정상 파싱
# --------------------------------------------------------------------------- #

def test_merged_title_above_header(tmp_path, engine):
    def build(ws):
        ws["A1"] = "방화벽 정책 신청서 (정보보호센터)"; ws.merge_cells("A1:L1")
        ws.merge_cells("C2:D2"); ws["C2"] = "출발지정보"
        ws.merge_cells("E2:F2"); ws["E2"] = "목적지정보"
        hdr = ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜",
               "포트", "방향", "용도", "시작일", "종료일", "비고"]
        for i, h in enumerate(hdr, 1):
            ws.cell(3, i, h)
        ws.cell(4, 1, 1); ws.cell(4, 2, "10.10.10.5"); ws.cell(4, 3, "PC")
        ws.cell(4, 4, "10.20.20.5"); ws.cell(4, 5, "DMZ"); ws.cell(4, 6, "TCP")
        ws.cell(4, 7, "443"); ws.cell(4, 8, "OUT"); ws.cell(4, 9, "업무")

    p = _build(tmp_path, "title.xlsx", build)
    parsed = parse_request_sheet(_rows_from(p))
    assert len(parsed) == 1
    assert parsed[0]["source_ip"] == "10.10.10.5"
    assert parsed[0]["dest_ip"] == "10.20.20.5"
    r = engine.analyze(parsed[0]["source_ip"], parsed[0]["dest_ip"], parsed[0]["direction"])
    assert r.status == "OK"


# --------------------------------------------------------------------------- #
# S3: 헤더 가로 병합 — 헤더 셀이 가로 병합돼도 좌상단 값으로 인식
# --------------------------------------------------------------------------- #

def test_horizontal_merge_in_header(tmp_path, engine):
    def build(ws):
        # 헤더 행에서 비고가 N1:O1로 가로 병합
        hdr = ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜",
               "포트", "방향", "용도", "시작일", "종료일", "비고"]
        for i, h in enumerate(hdr, 1):
            ws.cell(1, i, h)
        ws.merge_cells("L1:M1")  # 비고 병합 (빈 M 포함)
        ws.cell(2, 1, 1); ws.cell(2, 2, "10.10.10.5"); ws.cell(2, 3, "PC")
        ws.cell(2, 4, "10.20.20.5"); ws.cell(2, 5, "DMZ"); ws.cell(2, 6, "TCP")
        ws.cell(2, 7, "443"); ws.cell(2, 8, "OUT"); ws.cell(2, 9, "업무"); ws.cell(2, 12, "메모")

    p = _build(tmp_path, "hmerge.xlsx", build)
    parsed = parse_request_sheet(_rows_from(p))
    assert len(parsed) == 1
    assert parsed[0]["note"] == "메모"
    r = engine.analyze(parsed[0]["source_ip"], parsed[0]["dest_ip"], parsed[0]["direction"])
    assert r.status == "OK"


# --------------------------------------------------------------------------- #
# S4: 양쪽 IP 세로병합 + 아래 행은 포트만 다름 (Oracle이 지적한 last-row 누락 케이스)
# 한 신청이 동일 출발/목적 IP로 여러 포트를 신청하는 양식.
# --------------------------------------------------------------------------- #

def test_both_ip_vertical_merge_lower_row_kept(tmp_path, engine):
    def build(ws):
        hdr = ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜",
               "포트", "방향", "용도", "시작일", "종료일", "비고"]
        for i, h in enumerate(hdr, 1):
            ws.cell(1, i, h)
        # 출발지IP(B), 출발지(C), 목적지IP(D), 목적지(E) 모두 B2:E3 구간으로 세로병합,
        # 아래 행은 포트/용도만 다름.
        ws.cell(2, 2, "10.10.10.5"); ws.merge_cells("B2:B3")
        ws.cell(2, 3, "PC"); ws.merge_cells("C2:C3")
        ws.cell(2, 4, "10.20.20.5"); ws.merge_cells("D2:D3")
        ws.cell(2, 5, "DMZ"); ws.merge_cells("E2:E3")
        # row2: 포트 443 / row3: 포트 8443 (IP 컴럼은 row3에 원래 빈값)
        ws.cell(2, 1, 1); ws.cell(2, 6, "TCP"); ws.cell(2, 7, "443"); ws.cell(2, 8, "OUT"); ws.cell(2, 9, "웹")
        ws.cell(3, 1, 2); ws.cell(3, 6, "TCP"); ws.cell(3, 7, "8443"); ws.cell(3, 8, "OUT"); ws.cell(3, 9, "관리")

    p = _build(tmp_path, "bothmerge.xlsx", build)
    parsed = parse_request_sheet(_rows_from(p))
    # 두 행 모두 살아있어야 함 (last-row 누락 없음)
    assert len(parsed) == 2
    assert parsed[0]["source_ip"] == "10.10.10.5" and parsed[0]["dest_ip"] == "10.20.20.5"
    assert parsed[1]["source_ip"] == "10.10.10.5" and parsed[1]["dest_ip"] == "10.20.20.5"
    assert parsed[0]["port"] == "443" and parsed[1]["port"] == "8443"
    # 둘 다 동일 경로 (internal -> transit -> dmz)
    for r in parsed:
        res = engine.analyze(r["source_ip"], r["dest_ip"], r["direction"])
        assert res.status == "OK"
        assert res.target_firewalls == "SECUI-FW-01;SECUI-FW-02"
