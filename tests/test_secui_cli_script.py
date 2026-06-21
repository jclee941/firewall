from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl

from scripts.secui_cli_runtime import RequestRecord, secui_cli_rows
from scripts.workbook_contract import FIREWALLS, FIREWALL_RANGES, VENDOR_CLI_TEMPLATES

ROOT = Path(__file__).resolve().parents[1]
PY = sys.executable
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


def test_secui_cli_script_generates_text_from_workbook(tmp_path: Path) -> None:
    subprocess.run([PY, str(ROOT / "scripts" / "build_xlsm.py")], cwd=ROOT, check=True)
    output = tmp_path / "secui.txt"

    result = subprocess.run(
        [PY, str(ROOT / "scripts" / "secui_cli.py"), "--workbook", str(XLSM), "--output", str(output)],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    text = output.read_text(encoding="utf-8")
    assert "SECUI-FW-01" in text
    assert "fw set srule" in text
    assert "fw set addrgrp" in text


def test_secui_cli_script_generates_xlsx_from_request_folder(tmp_path: Path) -> None:
    request_folder = tmp_path / "request-folder"
    subprocess.run(
        [PY, str(ROOT / "scripts" / "make_request_folder.py"), "--output", str(request_folder)],
        cwd=ROOT,
        check=True,
    )
    text_output = tmp_path / "secui.txt"
    xlsx_output = tmp_path / "secui.xlsx"

    result = subprocess.run(
        [
            PY,
            str(ROOT / "scripts" / "secui_cli.py"),
            "--request-folder",
            str(request_folder),
            "--output",
            str(text_output),
            "--output-xlsx",
            str(xlsx_output),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    assert text_output.read_text(encoding="utf-8").count("fw set srule") >= 3
    wb = openpyxl.load_workbook(xlsx_output)
    try:
        assert "secui_cli" in wb.sheetnames
        assert wb["secui_cli"].max_row >= 4
    finally:
        wb.close()


def test_secui_cli_export_command_writes_xlsx_from_workbook(tmp_path: Path) -> None:
    subprocess.run([PY, str(ROOT / "scripts" / "build_xlsm.py")], cwd=ROOT, check=True)
    output = tmp_path / "secui-export.xlsx"

    result = subprocess.run(
        [
            PY,
            str(ROOT / "scripts" / "secui_cli.py"),
            "export",
            "--workbook",
            str(XLSM),
            "--format",
            "xlsx",
            "--output",
            str(output),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(output)
    try:
        ws = wb["secui_cli"]
        commands = [str(row[3].value or "") for row in ws.iter_rows(min_row=2)]
        assert any("fw set srule" in command for command in commands)
        assert any("fw set addrgrp" in command for command in commands)
    finally:
        wb.close()


def test_secui_cli_rows_derive_target_firewalls_when_parsed_request_has_blank_target() -> None:
    requests: list[RequestRecord] = [
        {
            "요청부서": "정보보호센터",
            "요청번호": "AUTO",
            "원본파일": "blank-target.xlsx",
            "원본행": 2,
            "대상방화벽": "",
            "출발지IP": "10.10.10.5",
            "출발지설명": "업무PC",
            "목적지IP": "10.20.20.5",
            "목적지설명": "DMZ서버",
            "프로토콜": "TCP",
            "포트": "443",
            "방향": "OUT",
            "용도": "자동 경로 탐색",
            "비고": "대상방화벽 공란",
        },
    ]

    rows = secui_cli_rows(requests, FIREWALLS, FIREWALL_RANGES, VENDOR_CLI_TEMPLATES)

    firewalls = [str(row[1]) for row in rows[1:]]
    assert firewalls == ["SECUI-FW-01", "SECUI-FW-02"]
    commands = "\n".join(str(row[3]) for row in rows[1:])
    assert "# device=SECUI-FW-01" in commands
    assert "# device=SECUI-FW-02" in commands
    assert "fw set srule" in commands
