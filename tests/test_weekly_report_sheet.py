from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


@pytest.fixture(scope="module")
def xlsm_path() -> Path:
    _ = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "build_xlsm.py")],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    assert XLSM.exists(), "builder did not produce the .xlsm"
    return XLSM


def test_weekly_report_sheet_describes_work_improvement_in_progress(xlsm_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    try:
        ws = wb["주간보고"]
        sheet_state = ws.sheet_state
        header = (ws.cell(1, 1).value, ws.cell(1, 2).value)
        values = [
            str(ws.cell(row=row, column=column).value or "")
            for row in range(1, ws.max_row + 1)
            for column in range(1, ws.max_column + 1)
        ]
        text = "\n".join(values)
    finally:
        wb.close()

    assert sheet_state == "visible"
    assert header == ("항목", "내용")
    assert "업무개선" in text
    assert "개발 진행" in text
    assert "Windows Excel 환경" in text
    assert "공개 준비" not in text
