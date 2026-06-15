from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


def _build() -> None:
    _ = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "build_xlsm.py")],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )


def _usage_links(ws: Worksheet) -> set[str]:
    links: set[str] = set()
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.hyperlink is None or cell.hyperlink.target is None:
            continue
        links.add(str(cell.value))
    return links


def test_cli_only_workbook_omits_policy_review_surfaces() -> None:
    _build()
    wb = openpyxl.load_workbook(XLSM, keep_vba=True, data_only=False)
    try:
        assert {"secui_policy_export", "policy_analysis", "policy_summary"}.isdisjoint(wb.sheetnames)
        assert "secui_cli" in _usage_links(wb["usage"])
    finally:
        wb.close()


def test_batch_surface_is_not_part_of_cli_only_workbook() -> None:
    _build()
    wb = openpyxl.load_workbook(XLSM, keep_vba=True, data_only=False)
    try:
        assert "secui_batch" not in wb.sheetnames
        assert "secui_batch" not in _usage_links(wb["usage"])
    finally:
        wb.close()


def test_convenience_surfaces_keep_core_contracts() -> None:
    _build()
    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    try:
        assert wb["requests"].max_column == 25
        assert [wb["firewall_ranges"].cell(1, column).value for column in range(1, 8)] == [
            "firewall_name",
            "source_cidr",
            "destination_cidr",
            "direction",
            "path_order",
            "enabled",
            "comment",
        ]
    finally:
        wb.close()
