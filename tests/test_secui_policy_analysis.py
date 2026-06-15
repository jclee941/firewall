import os
import subprocess
import sys

import openpyxl


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSM = os.path.join(ROOT, "dist", "firewall-policy-automation.xlsm")
PY = sys.executable
VBA_POLICY = os.path.join(ROOT, "vba", "FirewallPolicyAutomation.bas")

LEGACY_POLICY_SHEETS = {
    "secui_policy_export",
    "policy_analysis",
    "policy_summary",
}


def _build() -> None:
    _ = subprocess.run(
        [PY, os.path.join(ROOT, "scripts", "build_xlsm.py")],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )


def test_legacy_secui_policy_analysis_sheets_are_not_built() -> None:
    _build()
    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    try:
        assert LEGACY_POLICY_SHEETS.isdisjoint(wb.sheetnames)
        assert {"requests", "secui_cli", "vendor_cli_templates"} <= set(wb.sheetnames)
    finally:
        wb.close()


def test_legacy_secui_policy_analysis_vba_is_not_shipped() -> None:
    src = open(VBA_POLICY, encoding="utf-8").read()

    assert "AnalyzeSecuiPolicyExport" not in src
    assert "WriteSecuiPolicyExportHeaders" not in src
    assert "WritePolicyAnalysisHeaders" not in src
    assert "WritePolicySummarySheet" not in src
