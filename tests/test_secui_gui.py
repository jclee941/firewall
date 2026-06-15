from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from firewall_policy.gui_export import ExportRequest, build_export_command, run_export_request

ROOT = Path(__file__).resolve().parents[1]
PY = sys.executable
CLI = ROOT / "scripts" / "secui_cli.py"
XLSM = ROOT / "dist" / "firewall-policy-automation.xlsm"


def test_gui_export_command_targets_workbook_xlsx_output(tmp_path: Path) -> None:
    output = tmp_path / "secui.xlsx"
    command = build_export_command(
        PY,
        CLI,
        ExportRequest(
            source_mode="workbook",
            output_format="xlsx",
            workbook=XLSM,
            xlsx_output=output,
        ),
    )

    assert command == [
        PY,
        str(CLI),
        "export",
        "--workbook",
        str(XLSM),
        "--format",
        "xlsx",
        "--output-xlsx",
        str(output),
    ]


def test_gui_export_request_requires_matching_outputs(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="both export requires"):
        build_export_command(
            PY,
            CLI,
            ExportRequest(
                source_mode="request_folder",
                output_format="both",
                request_folder=tmp_path,
                text_output=tmp_path / "secui.txt",
            ),
        )


def test_gui_export_command_runs_existing_cli_surface(tmp_path: Path) -> None:
    subprocess.run([PY, str(ROOT / "scripts" / "build_xlsm.py")], cwd=ROOT, check=True)
    output = tmp_path / "secui.xlsx"
    command = build_export_command(
        PY,
        CLI,
        ExportRequest(
            source_mode="workbook",
            output_format="xlsx",
            workbook=XLSM,
            xlsx_output=output,
        ),
    )

    result = subprocess.run(command, cwd=ROOT, capture_output=True, text=True)

    assert result.returncode == 0, result.stderr
    wb = openpyxl.load_workbook(output)
    try:
        ws = wb["secui_cli"]
        commands = [str(row[3].value or "") for row in ws.iter_rows(min_row=2)]
        assert any("fw set srule" in command for command in commands)
    finally:
        wb.close()


def test_gui_script_smoke_creates_native_window() -> None:
    pytest.importorskip("PySide6")
    result = subprocess.run(
        [PY, str(ROOT / "scripts" / "secui_gui.py"), "--smoke"],
        cwd=ROOT,
        env={**os.environ, "QT_QPA_PLATFORM": "offscreen"},
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    assert "SECUI_CLI_GUI_SMOKE_OK" in result.stdout


def test_gui_export_request_runs_without_python_subprocess(tmp_path: Path) -> None:
    subprocess.run([PY, str(ROOT / "scripts" / "build_xlsm.py")], cwd=ROOT, check=True)
    output = tmp_path / "secui-direct.xlsx"

    run_export_request(
        ExportRequest(
            source_mode="workbook",
            output_format="xlsx",
            workbook=XLSM,
            xlsx_output=output,
        )
    )

    wb = openpyxl.load_workbook(output)
    try:
        ws = wb["secui_cli"]
        assert ws.max_row >= 4
    finally:
        wb.close()
