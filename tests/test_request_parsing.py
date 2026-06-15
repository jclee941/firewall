"""End-to-end 신청서 엑셀 파싱 검증.

Builds REAL .xlsx request files with varied layouts (column order, alias
headers, No on column B, header not on row 1), parses them with the VBA-mirror
parser, then routes each parsed row through the route engine — proving the full
pipeline (엑셀 파싱 -> zone 해석 -> 경로 산정) works on realistic forms.

Run: .venv/bin/python -m pytest tests/test_request_parsing.py -v
"""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(__file__))

from request_parser_oracle import (  # noqa: E402
    RequestParseError,
    canonical_header_name,
    find_header_row,
    find_request_sheet,
    header_key,
    parse_request_sheet,
    parse_request_sheet_exploded,
    explode_request_row,
    split_list,
    _norm_list,
    select_request_sheet,
)
from route_oracle import (  # noqa: E402
    Firewall,
    FirewallRange,
    RouteEngine,
)


# --------------------------------------------------------------------------- #
# Shared topology for routing the parsed rows (matches shipped seed semantics)
# --------------------------------------------------------------------------- #

@pytest.fixture
def engine():
    fws = [Firewall("SECUI-FW-01"), Firewall("SECUI-FW-02"), Firewall("SECUI-FW-03")]
    ranges = [
        FirewallRange("SECUI-FW-01", "10.10.0.0/16", "172.16.0.0/16", "OUT", 10),
        FirewallRange("SECUI-FW-01", "10.10.0.0/16", "10.20.0.0/16", "OUT", 10),
        FirewallRange("SECUI-FW-02", "10.10.0.0/16", "10.20.0.0/16", "OUT", 20),
        FirewallRange("SECUI-FW-01", "10.10.0.0/16", "8.8.8.0/24", "OUT", 10),
        FirewallRange("SECUI-FW-02", "10.10.0.0/16", "8.8.8.0/24", "OUT", 20),
        FirewallRange("SECUI-FW-03", "10.10.0.0/16", "8.8.8.0/24", "OUT", 30),
    ]
    return RouteEngine(firewalls=fws, firewall_ranges=ranges)


def _sheet_rows(ws):
    """Read an openpyxl worksheet into list[list] (the parser's input shape)."""
    return [[c.value for c in row] for row in ws.iter_rows()]


def _build_xlsx(tmp_path, name, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    p = tmp_path / name
    wb.save(p)
    return p


# --------------------------------------------------------------------------- #
# Unit: header normalization + alias + header-row detection
# --------------------------------------------------------------------------- #

def test_header_key_normalizes():
    assert header_key(" 출발지 IP ") == "출발지ip"
    assert header_key("Src IP") == "srcip"
    assert header_key(None) == ""


def test_canonical_aliases():
    assert canonical_header_name("srcip") == "출발지ip"
    assert canonical_header_name("source") == "출발지ip"
    assert canonical_header_name("출발ip") == "출발지ip"
    assert canonical_header_name("dstip") == "목적지ip"
    assert canonical_header_name("destination") == "목적지ip"
    assert canonical_header_name("destiation") == "목적지ip"
    assert canonical_header_name("protocol") == "프로토콜"
    assert canonical_header_name("목적지포트") == "포트"
    assert canonical_header_name("unknowncol") == "unknowncol"


def test_find_header_row_scans_top_30():
    rows = [
        ["방화벽 정책 신청서"],          # row 1 title
        [],                              # row 2 blank
        ["No", "출발지IP", "목적지IP"],   # row 3 header
    ]
    assert find_header_row(rows) == 3


def test_find_header_row_detects_by_header_content_without_no():
    """Header row is found by its FIELD HEADERS (출발지IP/목적지IP), even with no
    No/번호 column at all — the 'No' anchor is optional, not required."""
    rows = [["출발지IP", "목적지IP", "프로토콜", "포트"]]
    assert find_header_row(rows) == 1


def test_no_label_variants_canonicalize_to_no():
    """Each No-column label must canonicalize to 'no' BY ITSELF (not merely be
    tolerated because sibling field headers exist). Guards the false-positive
    where '#' was silently dropped yet the test still passed via other columns."""
    for no_label in ("No", "No.", "\ubc88\ud638", "\uc21c\ubc88", "\uc5f0\ubc88", "\uc21c \ubc88", "Seq", "#"):
        assert canonical_header_name(header_key(no_label)) == "no", \
            f"{no_label!r} does not canonicalize to 'no'"


def test_find_header_row_no_variants_detected():
    """No-column spelling variants must all anchor the header row: No. 순번 연번
    '순 번' Seq # — previously only exact 'no'/'번호' worked."""
    for no_label in ("No.", "순번", "연번", "순 번", "Seq", "#"):
        rows = [[no_label, "출발지IP", "목적지IP", "프로토콜"]]
        assert find_header_row(rows) == 1, f"{no_label!r} not detected"


def test_no_label_anchors_header_via_has_no_only():
    """A No-label + ONE IP column (no other field header) must still anchor the
    header row through the has_no path — proves the No-label itself is recognized,
    since field_count would be only 1 here."""
    for no_label in ("No", "No.", "\uc21c\ubc88", "#"):
        rows = [[no_label, "\ucd9c\ubc1c\uc9c0IP"]]  # has_no + has_ip, field_count==1
        assert find_header_row(rows) == 1, f"{no_label!r} did not anchor via has_no"


def test_find_header_row_picks_best_scoring_row():
    """When several rows have some text, the row with the MOST recognized field
    headers wins (and must contain a required IP column)."""
    rows = [
        ["방화벽 정책 신청서", None, None],        # title, 0 field headers
        ["작성자", "홍길동", "부서", "인프라"],     # metadata, 0 field headers
        ["No", "출발지IP", "목적지IP", "프로토콜", "포트", "방향"],  # real header
    ]
    assert find_header_row(rows) == 3


def test_find_header_row_truly_missing_raises():
    """A sheet with no recognizable field headers anywhere still raises."""
    rows = [["제목", "내용"], ["가", "나"]]
    with pytest.raises(RequestParseError):
        find_header_row(rows)


def test_find_header_row_requires_ip_column():
    """A row with field-ish headers but NO IP column is not a valid header row."""
    rows = [["프로토콜", "포트", "방향", "용도"]]  # no 출발지IP/목적지IP
    with pytest.raises(RequestParseError):
        find_header_row(rows)


def test_find_header_row_uses_user_aliases():
    """Header detection must honor settings header_alias so a fully custom IP
    column name still anchors the header row (parity with VBA UserAliasCanonical
    in FindHeaderRow). Without this, Python would raise where VBA succeeds."""
    # custom company headers that are NOT built-in aliases
    rows = [["Source Addr", "Dest Addr", "Service", "Port"]]
    # built-in fails (no recognized IP column)
    with pytest.raises(RequestParseError):
        find_header_row(rows)
    # with user aliases mapping the normalized keys to canonical IP names, it works
    aliases = {"sourceaddr": "\ucd9c\ubc1c\uc9c0ip", "destaddr": "\ubaa9\uc801\uc9c0ip"}
    assert find_header_row(rows, aliases) == 1


def test_parse_request_sheet_with_alias_only_ip_headers(tmp_path):
    """Public entry point: parse_request_sheet succeeds end-to-end when the IP
    columns are named ONLY by user aliases (no built-in match), exercising the
    full header-detect -> column-map -> extract path with user_aliases threaded
    through. Mirrors VBA where mUserAliases feeds both FindHeaderRow and
    BuildHeaderMap."""
    rows = [
        ["Source Addr", "Dest Addr", "Service", "Port"],
        ["10.10.10.5", "172.16.1.10", "TCP", "443"],
    ]
    p = _build_xlsx(tmp_path, "alias_ip.xlsx", rows)
    aliases = {"sourceaddr": "\ucd9c\ubc1c\uc9c0ip", "destaddr": "\ubaa9\uc801\uc9c0ip"}
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active),
                                 aliases)
    assert len(parsed) == 1
    assert parsed[0]["source_ip"] == "10.10.10.5"
    assert parsed[0]["dest_ip"] == "172.16.1.10"


# --------------------------------------------------------------------------- #
# E2E #1: canonical layout, header on row 1
# --------------------------------------------------------------------------- #

def test_parse_canonical_layout(tmp_path, engine):
    rows = [
        ["No", "대상방화벽", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜", "포트",
         "방향", "용도", "시작일", "종료일", "비고"],
        [1, "SECUI-FW-01", "10.10.10.5", "업무PC", "172.16.1.10", "서버", "TCP", "443",
         "OUT", "업무", "2026-01-01", "2026-12-31", "정기"],
    ]
    p = _build_xlsx(tmp_path, "req1.xlsx", rows)
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    assert len(parsed) == 1
    r = parsed[0]
    assert r["source_ip"] == "10.10.10.5"
    assert r["target_firewalls"] == "SECUI-FW-01"
    assert r["dest_ip"] == "172.16.1.10"
    assert r["protocol"] == "TCP"
    res = engine.analyze(r["source_ip"], r["dest_ip"], r["direction"])
    assert res.status == "OK"
    assert res.target_firewalls == "SECUI-FW-01"


# --------------------------------------------------------------------------- #
# E2E #2: alias headers + different column order + No on column B + title rows
# --------------------------------------------------------------------------- #

def test_parse_varied_layout(tmp_path, engine):
    rows = [
        ["방화벽 정책 신청서", None, None, None, None],       # row 1 title
        ["작성자: 홍길동", None, None, None, None],            # row 2 meta
        # row 3 header: No in column B, alias names, shuffled order
        [None, "No", "Dst IP", "Src IP", "Protocol", "Port", "Direction",
         "Usage", "Start", "End", "Note", "출발지", "목적지"],
        [None, 1, "10.20.20.5", "10.10.10.5", "tcp", "443", "OUT",
         "DMZ 연동", "2026-01-01", "2026-12-31", "비고1", "PC", "DMZ"],
        [None, 2, "172.16.1.10", "10.10.10.9", "udp", "53", "OUT",
         "DNS", "2026-01-01", "2026-12-31", "비고2", "PC", "DNS"],
    ]
    p = _build_xlsx(tmp_path, "req2.xlsx", rows)
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    assert len(parsed) == 2

    # row 1: 10.10.10.5 (internal) -> 10.20.20.5 (dmz) => FW-01;FW-02
    res1 = engine.analyze(parsed[0]["source_ip"], parsed[0]["dest_ip"], parsed[0]["direction"])
    assert res1.status == "OK"
    assert res1.target_firewalls == "SECUI-FW-01;SECUI-FW-02"
    assert res1.zone_path == "10.10.0.0/16>10.20.0.0/16"

    # row 2: 10.10.10.9 (internal) -> 172.16.1.10 (server) => FW-01
    res2 = engine.analyze(parsed[1]["source_ip"], parsed[1]["dest_ip"], parsed[1]["direction"])
    assert res2.status == "OK"
    assert res2.target_firewalls == "SECUI-FW-01"
    # protocol uppercased by parser
    assert parsed[0]["protocol"] == "TCP"
    assert parsed[1]["protocol"] == "UDP"


def test_parse_plain_english_source_destination_headers(tmp_path):
    rows = [
        ["No", "Source", "Destiation", "Protocol", "Port"],
        [1, "10.10.10.5", "10.20.20.5", "tcp", "443"],
    ]
    p = _build_xlsx(tmp_path, "english-source-destination.xlsx", rows)

    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))

    assert parsed[0]["source_ip"] == "10.10.10.5"
    assert parsed[0]["dest_ip"] == "10.20.20.5"


# --------------------------------------------------------------------------- #
# E2E #3: CIDR + address-list request values survive parsing and route
# --------------------------------------------------------------------------- #

def test_parse_cidr_and_list(tmp_path, engine):
    rows = [
        ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜", "포트",
         "방향", "용도", "시작일", "종료일", "비고"],
        [1, "10.10.10.0/24", "PC대역", "10.20.20.0/24", "DMZ대역", "TCP", "443",
         "OUT", "대역 신청", "2026-01-01", "2026-12-31", ""],
        [2, "10.10.10.5;10.10.10.6", "PC들", "172.16.1.10", "서버", "TCP", "22",
         "OUT", "리스트 신청", "2026-01-01", "2026-12-31", ""],
    ]
    p = _build_xlsx(tmp_path, "req3.xlsx", rows)
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    assert len(parsed) == 2

    r1 = engine.analyze(parsed[0]["source_ip"], parsed[0]["dest_ip"], parsed[0]["direction"])
    assert r1.status == "OK"
    assert r1.target_firewalls == "SECUI-FW-01;SECUI-FW-02"

    r2 = engine.analyze(parsed[1]["source_ip"], parsed[1]["dest_ip"], parsed[1]["direction"])
    assert r2.status == "OK"
    assert r2.target_firewalls == "SECUI-FW-01"


# --------------------------------------------------------------------------- #
# E2E #4: optional columns may be omitted; only the IP columns are required
# --------------------------------------------------------------------------- #

def test_parse_optional_columns_omitted_ok(tmp_path):
    # 비고/용도/날짜 등 omitted -> still parses (those are optional metadata)
    rows = [
        ["No", "\ucd9c\ubc1c\uc9c0IP", "\ubaa9\uc801\uc9c0IP", "\ud3ec\ud2b8"],
        [1, "10.10.10.5", "172.16.1.10", "443"],
    ]
    p = _build_xlsx(tmp_path, "req4a.xlsx", rows)
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    assert len(parsed) == 1
    assert parsed[0]["source_ip"] == "10.10.10.5"
    assert parsed[0]["dest_ip"] == "172.16.1.10"
    assert parsed[0]["note"] == ""      # absent optional column -> blank
    assert parsed[0]["purpose"] == ""


def test_parse_missing_ip_column_raises(tmp_path):
    # omitting a truly-required IP column still raises
    rows = [
        ["No", "\ucd9c\ubc1c\uc9c0IP", "\ud504\ub85c\ud1a0\ucf5c", "\ud3ec\ud2b8"],   # \ubaa9\uc801\uc9c0IP omitted
        [1, "10.10.10.5", "TCP", "443"],
    ]
    p = _build_xlsx(tmp_path, "req4b.xlsx", rows)
    with pytest.raises(RequestParseError) as exc:
        parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    assert "\ubaa9\uc801\uc9c0ip" in str(exc.value)


# --------------------------------------------------------------------------- #
# E2E #5: blank data rows skipped; trailing blanks ignored
# --------------------------------------------------------------------------- #

def test_parse_skips_blank_rows(tmp_path, engine):
    rows = [
        ["No", "출발지IP", "출발지", "목적지IP", "목적지", "프로토콜", "포트",
         "방향", "용도", "시작일", "종료일", "비고"],
        [1, "10.10.10.5", "PC", "172.16.1.10", "서버", "TCP", "443",
         "OUT", "업무", "2026-01-01", "2026-12-31", ""],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        [3, "", "PC", "", "서버", "TCP", "443", "OUT", "빈IP", "", "", ""],
    ]
    p = _build_xlsx(tmp_path, "req5.xlsx", rows)
    parsed = parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))
    # only the one row with IP data is kept (row with empty src+dst IP skipped)
    assert len(parsed) == 1
    assert parsed[0]["source_ip"] == "10.10.10.5"


# --------------------------------------------------------------------------- #
# E2E #6: 비표준 헤더 + settings header_alias -> 파싱 -> 경로추적 (통합)
# --------------------------------------------------------------------------- #

def test_parse_with_user_aliases_then_route(tmp_path, engine):
    from user_alias_oracle import parse_user_aliases

    # settings header_alias: 비표준 헤더를 표준 컬럼으로 매핑
    user_aliases = parse_user_aliases(
        "출발지IP=출발지주소; 목적지IP=목적지주소; 출발지=출발지설명; "
        "목적지=목적지설명; 용도=신청사유"
    )
    # 신청서: 표준 별칭으로는 안 잡히는 헤더(출발지주소/목적지주소 등)를
    # settings header_alias 로 해석해야 파싱 성공.
    rows = [
        ["No", "출발지주소", "출발지설명", "목적지주소", "목적지설명",
         "프로토콜", "포트", "방향", "신청사유", "시작일", "종료일", "비고"],
        [1, "10.10.10.5", "PC", "10.20.20.5", "DMZ", "TCP", "443",
         "OUT", "DMZ 연동", "2026-01-01", "2026-12-31", "비고"],
    ]
    p = _build_xlsx(tmp_path, "req6.xlsx", rows)
    parsed = parse_request_sheet(
        _sheet_rows(openpyxl.load_workbook(p).active),
        user_aliases=user_aliases,
    )
    assert len(parsed) == 1
    r = parsed[0]
    assert r["source_ip"] == "10.10.10.5"
    assert r["dest_ip"] == "10.20.20.5"
    assert r["purpose"] == "DMZ 연동"
    res = engine.analyze(r["source_ip"], r["dest_ip"], r["direction"])
    assert res.status == "OK"
    assert res.target_firewalls == "SECUI-FW-01;SECUI-FW-02"
    assert res.zone_path == "10.10.0.0/16>10.20.0.0/16"


def test_user_alias_missing_without_setting_raises(tmp_path):
    # If the IP columns use names the built-in map does NOT know and no alias is
    # registered, the required IP columns are missing -> raises.
    rows = [
        ["No", "zzsrc", "\ucd9c\ubc1c\uc9c0", "zzdst", "\ubaa9\uc801\uc9c0",
         "\ud504\ub85c\ud1a0\ucf5c", "\ud3ec\ud2b8", "\ubc29\ud5a5", "\uc6a9\ub3c4", "\uc2dc\uc791\uc77c", "\uc885\ub8cc\uc77c", "\ube44\uace0"],
        [1, "10.10.10.5", "PC", "10.20.20.5", "DMZ", "TCP", "443",
         "OUT", "\uc5f0\ub3d9", "2026-01-01", "2026-12-31", ""],
    ]
    p = _build_xlsx(tmp_path, "req7.xlsx", rows)
    with pytest.raises(RequestParseError):
        parse_request_sheet(_sheet_rows(openpyxl.load_workbook(p).active))


# --------------------------------------------------------------------------- #
# E2E #8: 시트 자동탐지 — 데이터가 첫 시트가 아닌 양식도 파싱
# --------------------------------------------------------------------------- #

def _build_multisheet_xlsx(tmp_path, name, named_sheets):
    """named_sheets: list of (title, rows). First entry becomes the active sheet."""
    wb = openpyxl.Workbook()
    first_title, first_rows = named_sheets[0]
    ws = wb.active
    ws.title = first_title
    for r in first_rows:
        ws.append(r)
    for title, rows in named_sheets[1:]:
        s = wb.create_sheet(title)
        for r in rows:
            s.append(r)
    p = tmp_path / name
    wb.save(p)
    return p


def _workbook_sheets(path):
    wb = openpyxl.load_workbook(path)
    return [[[c.value for c in row] for row in ws.iter_rows()] for ws in wb.worksheets]


_REQUEST_ROWS = [
    ["No", "출발지IP", "출발지", "목적지IP", "목적지",
     "프로토콜", "포트", "방향", "용도", "시작일", "종료일", "비고"],
    [1, "10.10.10.5", "PC", "10.20.20.5", "DMZ", "TCP", "443",
     "OUT", "연동", "2026-01-01", "2026-12-31", ""],
]


def test_find_request_sheet_picks_data_sheet_not_first():
    # 첫 시트는 결재/안내문(헤더 없음), 둘째 시트에 실제 신청 테이블이 있다.
    cover = [["방화벽 정책 신청서"], ["결재라인"], [""]]
    sheets = [cover, _REQUEST_ROWS]
    assert find_request_sheet(sheets) == 1


def test_find_request_sheet_ties_keep_leftmost():
    # 두 시트가 동일 점수면 왼쪽(첫) 시트가 이긴다 (기존 동작 보존).
    sheets = [_REQUEST_ROWS, _REQUEST_ROWS]
    assert find_request_sheet(sheets) == 0


def test_find_request_sheet_raises_when_no_sheet_has_header():
    sheets = [[["제목"], ["메모"]], [["a", "b"], [1, 2]]]
    with pytest.raises(RequestParseError):
        find_request_sheet(sheets)


def test_multisheet_xlsx_data_on_second_sheet_parses(tmp_path, engine):
    p = _build_multisheet_xlsx(
        tmp_path, "req8.xlsx",
        [("결재", [["방화벽 정책 신청서"], ["결재라인"]]),
         ("신청내역", _REQUEST_ROWS)],
    )
    sheets = _workbook_sheets(p)
    idx = find_request_sheet(sheets)
    assert idx == 1
    parsed = parse_request_sheet(sheets[idx])
    assert len(parsed) == 1
    r = parsed[0]
    assert r["source_ip"] == "10.10.10.5"
    assert r["dest_ip"] == "10.20.20.5"
    res = engine.analyze(r["source_ip"], r["dest_ip"], r["direction"])
    assert res.status == "OK"


# --------------------------------------------------------------------------- #
# E2E #9: 명시적 파싱 대상 시트 선택 (settings.parse_sheet)
# --------------------------------------------------------------------------- #

_HIGH_SCORE_ROWS = [
    ["No", "출발지IP", "출발지", "목적지IP", "목적지",
     "프로토콜", "포트", "방향", "용도", "시작일", "종료일", "비고"],
    [1, "10.10.10.9", "PC", "10.20.20.9", "SRV", "TCP", "80",
     "OUT", "오답", "2026-01-01", "2026-12-31", ""],
]


def test_select_request_sheet_explicit_name_wins_over_auto_detect():
    # '고득점' 시트가 더 높은 점수 + 좌측이지만, parse_sheet로 '신청내역' 명시 선택.
    named = [
        ("고득점", _HIGH_SCORE_ROWS),
        ("신청내역", _REQUEST_ROWS),
        ("표지", [["방화벽 신청서"]]),
    ]
    # auto-detect는 좌측(고득점, idx 0)을 고르지만, 명시명은 idx 1.
    assert find_request_sheet([r for _n, r in named]) == 0
    assert select_request_sheet(named, "신청내역") == 1


def test_select_request_sheet_missing_name_raises_naming_it():
    named = [("신청내역", _REQUEST_ROWS)]
    with pytest.raises(RequestParseError) as ei:
        select_request_sheet(named, "없는시트")
    assert "없는시트" in str(ei.value)


def test_select_request_sheet_blank_falls_back_to_auto_detect():
    # 빈 parse_sheet -> auto-detect; 동점이면 좌측(첫) 시트.
    named = [("첫시트", _REQUEST_ROWS), ("둘째시트", _REQUEST_ROWS)]
    assert select_request_sheet(named, "") == 0
    # 데이터가 둘째 시트에만 있으면 auto-detect가 둘째를 고름.
    named2 = [("표지", [["제목"]]), ("신청내역", _REQUEST_ROWS)]
    assert select_request_sheet(named2, "") == 1


def test_select_request_sheet_named_sheet_without_header_raises():
    # 명시한 시트가 존재하지만 헤더가 없으면 fallback 없이 에러.
    named = [
        ("표지", [["방화벽 신청서"], ["결재라인"]]),
        ("신청내역", _REQUEST_ROWS),
    ]
    with pytest.raises(RequestParseError) as ei:
        select_request_sheet(named, "표지")
    assert "표지" in str(ei.value)


def test_select_request_sheet_e2e_named_xlsx(tmp_path, engine):
    p = _build_multisheet_xlsx(
        tmp_path, "req9.xlsx",
        [("고득점", _HIGH_SCORE_ROWS),
         ("신청내역", _REQUEST_ROWS)],
    )
    wb = openpyxl.load_workbook(p)
    named = [(ws.title, [[c.value for c in row] for row in ws.iter_rows()])
             for ws in wb.worksheets]
    idx = select_request_sheet(named, "신청내역")
    assert idx == 1
    parsed = parse_request_sheet(named[idx][1])
    assert len(parsed) == 1
    assert parsed[0]["source_ip"] == "10.10.10.5"


# --------------------------------------------------------------------------- #
# E2E #10: cartesian explode — multi-valued 출발지IP/목적지IP/포트 → N rules
# --------------------------------------------------------------------------- #

def test_explode_full_cartesian_product(tmp_path):
    rows = [
        ["No", "\ucd9c\ubc1c\uc9c0IP", "\ucd9c\ubc1c\uc9c0", "\ubaa9\uc801\uc9c0IP", "\ubaa9\uc801\uc9c0", "\ud504\ub85c\ud1a0\ucf5c", "\ud3ec\ud2b8",
         "\ubc29\ud5a5", "\uc6a9\ub3c4", "\uc2dc\uc791\uc77c", "\uc885\ub8cc\uc77c", "\ube44\uace0"],
        [1, "10.0.0.1;10.0.0.2", "PC", "20.0.0.1;20.0.0.2", "SRV", "TCP", "80;443",
         "OUT", "\uc5f0\ub3d9", "2026-01-01", "2026-12-31", ""],
    ]
    p = _build_xlsx(tmp_path, "explode.xlsx", rows)
    exploded = parse_request_sheet_exploded(_sheet_rows(openpyxl.load_workbook(p).active))
    # 2 src × 2 dst × 2 port = 8 rules
    assert len(exploded) == 8
    # each exploded row holds SINGLE values
    for r in exploded:
        assert ";" not in r["source_ip"]
        assert ";" not in r["dest_ip"]
        assert ";" not in r["port"]
    # the full product is present
    triples = {(r["source_ip"], r["dest_ip"], r["port"]) for r in exploded}
    assert triples == {
        ("10.0.0.1", "20.0.0.1", "80"), ("10.0.0.1", "20.0.0.1", "443"),
        ("10.0.0.1", "20.0.0.2", "80"), ("10.0.0.1", "20.0.0.2", "443"),
        ("10.0.0.2", "20.0.0.1", "80"), ("10.0.0.2", "20.0.0.1", "443"),
        ("10.0.0.2", "20.0.0.2", "80"), ("10.0.0.2", "20.0.0.2", "443"),
    }
    # repeated metadata survives
    assert all(r["source_name"] == "PC" and r["protocol"] == "TCP" for r in exploded)


def test_explode_missing_port_keeps_one_row():
    # a field with no value yields exactly ONE blank element, row not dropped
    row = {"source_ip": "10.0.0.1", "dest_ip": "20.0.0.1", "port": "",
           "source_name": "PC", "protocol": "TCP"}
    out = explode_request_row(row)
    assert len(out) == 1
    assert out[0]["port"] == ""
    assert out[0]["source_ip"] == "10.0.0.1"


def test_split_list_handles_blanks_and_trailing():
    assert split_list("") == [""]
    assert split_list("80;443") == ["80", "443"]
    assert split_list(";80;;443;") == ["80", "443"]


def test_norm_list_preserves_tilde_and_trims_trailing_semicolon():
    # '~' (range marker) preserved; trailing ';' and spaces trimmed
    assert _norm_list("10.0.0.1~10.0.0.5") == "10.0.0.1~10.0.0.5"
    assert _norm_list(" 80 ;\n443 ; ") == "80;443"
    assert _norm_list("10.177.124.0/22~ ") == "10.177.124.0/22~"
