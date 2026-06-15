from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scripts.workbook_contract import (
    OPERATOR_VISIBLE_SHEETS,
    PROTECT_SHEETS,
    SUPPORT_DATA_SHEETS,
    TAB_COLORS,
)

UX_LAST_ROW = 5000
REQ_PROTOCOL_COL = 12
REQ_PORT_COL = 13
REQ_DIRECTION_COL = 14
REQ_TARGET_COL = 7
REQ_SRC_IP_COL = 8
REQ_DST_IP_COL = 10
REQ_HIDDEN_ROUTE_COLS = (6, 19, 20, 21, 22, 23, 24)
REQ_HEADER_ROW = 2
REQ_DATA_START_ROW = 3
EMPTY_REQUIRED_FILL = PatternFill("solid", fgColor="FFC7CE")
TARGET_FIREWALL_FILL = PatternFill("solid", fgColor="D9EAD3")
POLICY_ALLOW_FILL = PatternFill("solid", fgColor="D9EAD3")
POLICY_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
POLICY_BLOCK_FILL = PatternFill("solid", fgColor="F4CCCC")
LINK_FONT = Font(color="0563C1", underline="single")
EXPORT_HEADER_COMMENTS = {
    "A": "필수: SECUI export 정책 ID를 붙여넣습니다.",
    "B": "필수: 기존 정책명을 붙여넣습니다.",
    "C": "필수: firewalls 시트의 방화벽명과 같은 이름을 붙여넣습니다.",
    "D": "필수: 출발지 주소/CIDR/ANY 또는 객체명을 붙여넣습니다.",
    "E": "필수: 목적지 주소/CIDR/ANY 또는 객체명을 붙여넣습니다.",
    "F": "필수: tcp/443, udp/53, ANY 같은 서비스 표기를 붙여넣습니다.",
    "G": "필수: allow/deny/drop/reject/accept/pass 중 하나를 붙여넣습니다.",
    "H": "필수: Y/N 또는 TRUE/FALSE 사용여부를 붙여넣습니다.",
    "I": "선택: SECUI export 설명이나 운영 메모를 붙여넣습니다.",
}
CLI_TEMPLATE_HEADER_COMMENTS = {
    "A": "벤더명입니다. SECUI CLI 생성은 SECUI 행을 사용합니다.",
    "B": "템플릿 이름입니다. 운영자가 구분하기 위한 값입니다.",
    "C": "Y이면 사용합니다. 같은 벤더에서 처음 만나는 사용 템플릿을 적용합니다.",
    "D": "CLI 명령 템플릿입니다. {policy_name_q}, {source_interface_q}, {destination_interface_q}, {source_object_q}, {destination_object_q}, {service_object_q}, {description_q}, {firewall_name} 등을 사용할 수 있습니다.",
    "E": "secui_cli 검토메모 컬럼에 복사될 안내 문구입니다.",
}
USAGE_LINKS = (
    ("requests", "신청 통합/분석 결과"),
    ("firewalls", "방화벽 장비 목록"),
    ("settings", "신청서 폴더와 파싱 설정"),
    ("secui_batch", "SECUI 배치 입력용 출력"),
    ("secui_cli", "SECUI CLI 명령 초안"),
    ("secui_policy_export", "SECUI 기존 정책 export 붙여넣기"),
    ("policy_analysis", "기존 정책 분석 결과"),
    ("policy_summary", "기존 정책 분석 요약"),
    ("vendor_cli_templates", "SECUI CLI 템플릿 설정"),
)


def add_list_validation(ws, col_letter, values, *, allow_blank=True, start_row=2):
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "목록에서 값을 선택하세요."
    dv.errorTitle = "잘못된 입력"
    dv.prompt = "드롭다운에서 선택"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{UX_LAST_ROW}")
    return dv


def apply_ux(wb) -> None:
    for name, color in TAB_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    req = wb["requests"]
    for column in REQ_HIDDEN_ROUTE_COLS:
        req.column_dimensions[get_column_letter(column)].hidden = True
    add_list_validation(req, get_column_letter(REQ_PROTOCOL_COL), ["TCP", "UDP", "ICMP"], start_row=REQ_DATA_START_ROW)
    add_list_validation(req, get_column_letter(REQ_DIRECTION_COL), ["IN", "OUT", "BOTH"], start_row=REQ_DATA_START_ROW)

    if "sample-request-format" in wb.sheetnames:
        sf = wb["sample-request-format"]
        add_list_validation(sf, "H", ["TCP", "UDP", "ICMP"])
        add_list_validation(sf, "J", ["IN", "OUT", "BOTH"])

    if "usage" in wb.sheetnames:
        usage = wb["usage"]
        for row in range(1, usage.max_row + 1):
            for column in (3, 4):
                usage.cell(row, column).value = None
                usage.cell(row, column).hyperlink = None
        usage.cell(1, 3, "바로가기")
        usage.cell(1, 4, "설명")
        usage.column_dimensions["C"].width = 24
        usage.column_dimensions["D"].width = 34
        for row, (sheet_name, description) in enumerate(USAGE_LINKS, start=2):
            cell = usage.cell(row, 3, sheet_name)
            cell.hyperlink = f"#'{sheet_name}'!A1"
            cell.style = "Hyperlink"
            cell.font = LINK_FONT
            usage.cell(row, 4, description)
        wb.active = wb.sheetnames.index("usage")

    for name in OPERATOR_VISIBLE_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "visible"
    for name in SUPPORT_DATA_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"

    fw = wb["firewalls"]
    add_list_validation(fw, "C", ["Y", "N"])
    ranges = wb["firewall_ranges"]
    add_list_validation(ranges, "D", ["OUT", "IN", "BOTH"])
    add_list_validation(ranges, "F", ["Y", "N"])

    if "secui_policy_export" in wb.sheetnames:
        export = wb["secui_policy_export"]
        for column, text in EXPORT_HEADER_COMMENTS.items():
            export[f"{column}1"].comment = Comment(text, "firewall-automation")
        add_list_validation(export, "G", ["allow", "deny", "drop", "reject", "accept", "pass"])
        add_list_validation(export, "H", ["Y", "N", "TRUE", "FALSE"])

    if "vendor_cli_templates" in wb.sheetnames:
        templates = wb["vendor_cli_templates"]
        for column, text in CLI_TEMPLATE_HEADER_COMMENTS.items():
            templates[f"{column}1"].comment = Comment(text, "firewall-automation")
        add_list_validation(templates, "C", ["Y", "N"])

    if "policy_analysis" in wb.sheetnames:
        analysis = wb["policy_analysis"]
        for col in range(11, 21):
            analysis.column_dimensions[get_column_letter(col)].hidden = True
        status_rng = f"A2:A{UX_LAST_ROW}"
        analysis.conditional_formatting.add(
            status_rng,
            FormulaRule(formula=['$A2="EXISTING_ALLOW"'], fill=POLICY_ALLOW_FILL, stopIfTrue=False),
        )
        analysis.conditional_formatting.add(
            status_rng,
            FormulaRule(
                formula=['OR($A2="PARTIAL_MATCH",$A2="DISABLED_MATCH",$A2="OBJECT_UNRESOLVED")'],
                fill=POLICY_REVIEW_FILL,
                stopIfTrue=False,
            ),
        )
        analysis.conditional_formatting.add(
            status_rng,
            FormulaRule(
                formula=['OR($A2="EXISTING_DENY",$A2="NO_EXISTING_POLICY")'],
                fill=POLICY_BLOCK_FILL,
                stopIfTrue=False,
            ),
        )

    src_letter = get_column_letter(REQ_SRC_IP_COL)
    dst_letter = get_column_letter(REQ_DST_IP_COL)
    for letter in (src_letter, dst_letter):
        rng = f"{letter}{REQ_DATA_START_ROW}:{letter}{UX_LAST_ROW}"
        rule = FormulaRule(
            formula=[f"ISBLANK({letter}{REQ_DATA_START_ROW})"],
            fill=EMPTY_REQUIRED_FILL,
            stopIfTrue=False,
        )
        req.conditional_formatting.add(rng, rule)

    target_letter = get_column_letter(REQ_TARGET_COL)
    target_rng = f"{target_letter}{REQ_DATA_START_ROW}:{target_letter}{UX_LAST_ROW}"
    target_rule = FormulaRule(
        formula=[f"LEN(TRIM({target_letter}{REQ_DATA_START_ROW}))>0"],
        fill=TARGET_FIREWALL_FILL,
        stopIfTrue=False,
    )
    req.conditional_formatting.add(target_rng, target_rule)

    req.cell(REQ_HEADER_ROW, REQ_SRC_IP_COL).comment = Comment(
        "출발지 IP 또는 CIDR\n예: 10.10.10.0/24 또는 10.10.10.5\n여러 개는 ; 로 구분",
        "firewall-automation",
    )
    req.cell(REQ_HEADER_ROW, REQ_DST_IP_COL).comment = Comment(
        "목적지 IP 또는 CIDR\n예: 172.16.1.10 또는 172.16.1.0/24\n여러 개는 ; 로 구분",
        "firewall-automation",
    )
    req.cell(REQ_HEADER_ROW, REQ_PROTOCOL_COL).comment = Comment(
        "SECUI 서비스 표기용 프로토콜\n예: TCP, UDP, ICMP\nservice_catalog 시트의 secui_service 예시 참고",
        "firewall-automation",
    )
    req.cell(REQ_HEADER_ROW, REQ_PORT_COL).comment = Comment(
        "SECUI 서비스 표기용 포트\n예: 443 -> tcp/443, 53 -> udp/53\n목록에 없는 포트도 직접 입력 가능",
        "firewall-automation",
    )

    unlocked = Protection(locked=False)
    for name in PROTECT_SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        last_col = ws.max_column
        for row in range(2, UX_LAST_ROW + 1):
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).protection = unlocked
        ws.protection.enable()
        ws.protection.autoFilter = False
        ws.protection.sort = False
        ws.protection.formatCells = True
