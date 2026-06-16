from __future__ import annotations

import argparse
import sys
from collections.abc import Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Final

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from firewall_policy import request_parser
from firewall_policy.folder_parse import parse_request_folder_name
from firewall_policy.request_parser import (
    parse_request_sheet_exploded,
    select_request_sheet,
    sheet_to_filled_rows,
)
from scripts.secui_cli_runtime import Cell, RequestRecord, Table, secui_cli_rows
from scripts.workbook_contract import FIREWALLS, FIREWALL_RANGES, VENDOR_CLI_TEMPLATES

REQUEST_SHEET = "requests"
FIREWALLS_SHEET = "firewalls"
FIREWALL_RANGES_SHEET = "firewall_ranges"
TEMPLATE_SHEET = "vendor_cli_templates"
SETTINGS_SHEET = "settings"
HEADER_ALIASES_SHEET = "header_aliases"
WORKBOOK_REQUEST_HEADER_MAP: Final = {
    "출발지": "출발지IP",
    "목적지": "목적지IP",
}


def main() -> int:
    args = _parse_args()
    try:
        rows = _run(args)
    except FileNotFoundError as exc:
        print(f"ERROR: file not found: {exc.filename}", file=sys.stderr)
        return 2
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    try:
        _export_rows(args, rows)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate SECUI CLI without running Excel VBA macros.")
    parser.add_argument("command", nargs="?", choices=["export"], default="export", help="Export generated SECUI CLI.")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--workbook", help="Existing automation .xlsx/.xlsm workbook containing requests/config sheets.")
    source.add_argument("--request-folder", help="Folder tree containing request .xlsx files.")
    parser.add_argument("--config-workbook", help="Workbook supplying firewalls, firewall_ranges, and templates for folder mode.")
    parser.add_argument("--parse-sheet", default="", help="Exact request sheet name for folder workbooks; blank auto-detects.")
    parser.add_argument("--format", choices=["text", "xlsx", "both"], default="text", help="Export format.")
    parser.add_argument("--output", help="Write rendered SECUI CLI text to this file instead of stdout.")
    parser.add_argument("--output-xlsx", help="Also write a workbook with a secui_cli sheet.")
    return parser.parse_args()


def _run(args: argparse.Namespace) -> list[list[str | int]]:
    if args.workbook:
        workbook_path = Path(args.workbook)
        requests = _requests_from_workbook(workbook_path)
        firewalls = _table_from_workbook(workbook_path, FIREWALLS_SHEET, FIREWALLS)
        ranges = _table_from_workbook(workbook_path, FIREWALL_RANGES_SHEET, FIREWALL_RANGES)
        templates = _table_from_workbook(workbook_path, TEMPLATE_SHEET, VENDOR_CLI_TEMPLATES)
        return secui_cli_rows(requests, firewalls, ranges, templates)

    folder = Path(args.request_folder)
    if not folder.exists():
        raise FileNotFoundError(str(folder))
    config = Path(args.config_workbook) if args.config_workbook else None
    firewalls = _table_from_workbook(config, FIREWALLS_SHEET, FIREWALLS) if config else FIREWALLS
    ranges = _table_from_workbook(config, FIREWALL_RANGES_SHEET, FIREWALL_RANGES) if config else FIREWALL_RANGES
    templates = _table_from_workbook(config, TEMPLATE_SHEET, VENDOR_CLI_TEMPLATES) if config else VENDOR_CLI_TEMPLATES
    settings = _settings_from_workbook(config) if config else {}
    parse_sheet = str(args.parse_sheet or settings.get("parse_sheet", "") or "")
    aliases = _user_aliases_from_workbook(config, str(settings.get("header_alias", "") or "")) if config else {}
    requests = _requests_from_folder(folder, parse_sheet, aliases)
    return secui_cli_rows(requests, firewalls, ranges, templates)


def _requests_from_workbook(path: Path) -> list[RequestRecord]:
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        if REQUEST_SHEET not in wb.sheetnames:
            raise ValueError(f"missing sheet: {REQUEST_SHEET}")
        ws = wb[REQUEST_SHEET]
        headers = [
            _workbook_request_header(str(ws.cell(2, col).value or "").strip())
            for col in range(1, ws.max_column + 1)
        ]
        records: list[RequestRecord] = []
        for row_index in range(3, ws.max_row + 1):
            record: RequestRecord = {
                header: _cell_value(ws, row_index, col)
                for col, header in enumerate(headers, start=1)
                if header
            }
            if not str(record.get("출발지IP", "") or "").strip() and not str(record.get("목적지IP", "") or "").strip():
                continue
            record["원본행"] = record.get("원본행") or row_index
            records.extend(_explode_record(record))
        return records
    finally:
        wb.close()


def _workbook_request_header(header: str) -> str:
    return WORKBOOK_REQUEST_HEADER_MAP.get(header, header)


def _requests_from_folder(folder: Path, parse_sheet: str, user_aliases: dict[str, str] | None = None) -> list[RequestRecord]:
    records: list[RequestRecord] = []
    for path in sorted(folder.rglob("*.xlsx")):
        if path.name.startswith("~$") or "빈양식" in path.name:
            continue
        wb = openpyxl.load_workbook(path, data_only=False)
        try:
            named_rows = [(ws.title, sheet_to_filled_rows(ws, user_aliases)) for ws in wb.worksheets]
            sheet_index = select_request_sheet(named_rows, parse_sheet, user_aliases)
            parsed = parse_request_sheet_exploded(named_rows[sheet_index][1], user_aliases)
        finally:
            wb.close()
        team, doc_no, title = parse_request_folder_name(path.parent.name)
        for request in parsed:
            records.append(_record_from_parsed_request(request, team, doc_no, title, path.name))
    return records


def _record_from_parsed_request(
    request: Mapping[str, Cell],
    team: str,
    doc_no: str,
    title: str,
    source_file: str,
) -> RequestRecord:
    return {
        "요청부서": team,
        "요청번호": doc_no,
        "제목": title,
        "원본파일": source_file,
        "원본행": request.get("source_row", ""),
        "대상방화벽": request.get("target_firewalls", ""),
        "출발지IP": request.get("source_ip", ""),
        "출발지설명": request.get("source_name", ""),
        "목적지IP": request.get("dest_ip", ""),
        "목적지설명": request.get("dest_name", ""),
        "프로토콜": request.get("protocol", ""),
        "포트": request.get("port", ""),
        "방향": request.get("direction", ""),
        "용도": request.get("purpose", ""),
        "비고": request.get("note", ""),
    }


def _explode_record(record: RequestRecord) -> list[RequestRecord]:
    out: list[RequestRecord] = []
    for source in _split_list(str(record.get("출발지IP", "") or "")):
        for destination in _split_list(str(record.get("목적지IP", "") or "")):
            for port in _split_list(str(record.get("포트", "") or "")):
                item = dict(record)
                item["출발지IP"] = source
                item["목적지IP"] = destination
                item["포트"] = port
                out.append(item)
    return out


def _split_list(value: str) -> list[str]:
    text = value.strip(";")
    if not text:
        return [""]
    return [part for part in text.split(";") if part] or [""]


def _table_from_workbook(path: Path | None, sheet_name: str, fallback: Table) -> Table:
    if path is None:
        return fallback
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return fallback
        return _sheet_table(wb[sheet_name])
    finally:
        wb.close()


def _settings_from_workbook(path: Path) -> dict[str, str]:
    table = _table_from_workbook(path, SETTINGS_SHEET, [])
    settings: dict[str, str] = {}
    for row in table[1:]:
        if len(row) < 2:
            continue
        key = str(row[0] or "").strip().lower()
        if key:
            settings[key] = str(row[1] or "").strip()
    return settings


def _user_aliases_from_workbook(path: Path, header_alias: str) -> dict[str, str]:
    aliases = request_parser.parse_user_aliases(header_alias)
    table = _table_from_workbook(path, HEADER_ALIASES_SHEET, [])
    aliases.update(request_parser.aliases_from_rows(table[1:]))
    return aliases


def _sheet_table(ws: Worksheet) -> list[list[Cell]]:
    return [
        [_cell_value(ws, row, col) for col in range(1, ws.max_column + 1)]
        for row in range(1, ws.max_row + 1)
    ]


def _cell_value(ws: Worksheet, row: int, column: int) -> Cell:
    value = ws.cell(row=row, column=column).value
    if value is None or isinstance(value, str | int | float | bool | date | datetime):
        return value
    return str(value)


def _rows_to_text(rows: list[list[str | int]]) -> str:
    lines: list[str] = []
    for row in rows[1:]:
        lines.append(f"# {row[0]} {row[1]} {row[2]}")
        lines.append(str(row[3]))
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _export_rows(args: argparse.Namespace, rows: list[list[str | int]]) -> None:
    if args.format == "xlsx":
        xlsx_output = args.output_xlsx or args.output
        if not xlsx_output:
            raise ValueError("--format xlsx requires --output or --output-xlsx")
        _write_xlsx(Path(xlsx_output), rows)
        return

    if args.output:
        _write_text(Path(args.output), rows)
    else:
        print(_rows_to_text(rows))

    if args.format == "both" and not args.output_xlsx:
        raise ValueError("--format both requires --output-xlsx")
    if args.output_xlsx:
        _write_xlsx(Path(args.output_xlsx), rows)


def _write_text(path: Path, rows: list[list[str | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_rows_to_text(rows), encoding="utf-8")


def _write_xlsx(path: Path, rows: list[list[str | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("new output workbook did not create an active sheet")
    ws.title = "secui_cli"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    raise SystemExit(main())
