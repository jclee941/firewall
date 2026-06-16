#!/usr/bin/env python3
"""Build a REAL macro-enabled firewall-policy-automation.xlsm on Linux.

No Windows Excel / PowerShell / COM required. Uses:
  - pyOpenVBA  : inject the two VBA modules into a real vbaProject.bin
  - openpyxl   : pre-create and seed all worksheets while preserving the VBA

Run:
  ./.venv/bin/python scripts/build_xlsm.py
Output:
  dist/firewall-policy-automation.xlsm
"""

from __future__ import annotations

import os
import struct
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pyopenvba import ExcelFile

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.workbook_contract import (
    EXAMPLE_REQUEST_ROWS,
    FILTER_SHEETS,
    FIREWALLS,
    FIREWALL_RANGES,
    FREEZE_PANES,
    HEADER_ALIASES,
    PROCESSING_LOG,
    REQUEST_TRACKING_HEADERS,
    REQUEST_TRACKING_SHEET,
    REQUESTS_HEADERS,
    ROUTE_RESULTS_HEADERS,
    SAMPLE_FORMAT,
    SERVICE_CATALOG,
    SETTINGS,
    USAGE,
    VENDOR_CLI_TEMPLATES,
    WEEKLY_REPORT,
    WIDTHS,
)
from scripts.secui_cli_seed import secui_cli_seed_rows
from scripts.workbook_ux import apply_ux

VBA_DIR = ROOT / "vba"
DIST = ROOT / "dist"
OUT = DIST / "firewall-policy-automation.xlsm"

MODULES = [
    ("FirewallPolicyAutomation", VBA_DIR / "FirewallPolicyAutomation.bas"),
    ("FirewallRouteAnalysis", VBA_DIR / "FirewallRouteAnalysis.bas"),
]

# --------------------------------------------------------------------------- #
# Visual styling (first-open readability; VBA macros may re-format at runtime)
# --------------------------------------------------------------------------- #

_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_HEADER_FONT = Font(bold=True, color="1F3864")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BOTTOM = Border(bottom=Side(style="thin", color="9DB2CE"))

# requests output layout: row 1 = cosmetic group labels, row 2 = leaf headers,
# data from row 3. Mirror of VBA REQ_HEADER_GROUP_ROW / REQ_HEADER_ROW / REQ_DATA_START_ROW.
_REQ_HEADER_GROUP_ROW = 1
_REQ_HEADER_ROW = 2
_REQ_DATA_START_ROW = 3


def _style_sheet(ws, header_row: int = 1):
    from openpyxl.utils import get_column_letter
    for col_letter, width in WIDTHS.get(ws.title, {}).items():
        ws.column_dimensions[col_letter].width = width
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if cell.value is None:
            continue
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BOTTOM
    ws.freeze_panes = FREEZE_PANES.get(ws.title, "A2")
    if ws.title in FILTER_SHEETS:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(last_col)}"
            f"{max(ws.max_row, header_row)}"
        )


def _write_rows(ws, rows):
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

_AUTO_RUN_BODY = (
    "\r\n"
    "Private Sub Workbook_Open()\r\n"
    "    ' Auto-run on open: integrate the request folder and refresh SECUI outputs.\r\n"
    "    On Error GoTo AutoRunErr\r\n"
    "    AutoRunWorkbookOutputs\r\n"
    "    Exit Sub\r\n"
    "AutoRunErr:\r\n"
    "    MsgBox \"\uc790\ub3d9 \ucd9c\ub825 \uc0dd\uc131 \uc911 \uc624\ub958: \" & Err.Description, vbExclamation\r\n"
    "End Sub\r\n"
)


def _inject_auto_run(proj) -> None:
    """Append Workbook_Open to the ThisWorkbook document module so the workbook
    auto-integrates the request folder when opened (macros must be enabled)."""
    m = proj.get_module("ThisWorkbook")
    if "Workbook_Open" in m.source:
        return
    m.source = m.source.rstrip("\r\n") + _AUTO_RUN_BODY
    m.dirty = True


def _force_codepage_949(proj) -> None:
    """Patch the parsed VBA project so it serializes with PROJECTCODEPAGE=949.

    pyOpenVBA's serialize_dir_stream reuses the PROJECTINFORMATION prefix from
    proj.dir_raw verbatim (which carries PROJECTCODEPAGE=1252 from the create_new
    template), and rebuild_module_stream MBCS-encodes each module with
    proj.code_page. Setting proj.code_page alone is ignored because the dir prefix
    is spliced unchanged. So we patch the 0x0003 (PROJECTCODEPAGE) record inside
    dir_raw to 949 AND set proj.code_page=949, then mark dir + modules dirty so the
    streams are actually re-encoded. cp949 covers 한글, so the source survives."""
    raw = bytearray(proj.dir_raw)
    i = 0
    patched = False
    while i + 6 <= len(raw):
        rid = struct.unpack_from("<H", raw, i)[0]
        size = struct.unpack_from("<I", raw, i + 2)[0]
        if rid == 0x0003 and size == 2:  # PROJECTCODEPAGE
            struct.pack_into("<H", raw, i + 6, 949)
            patched = True
            break
        i += 6 + size
    if not patched:
        raise RuntimeError("PROJECTCODEPAGE record not found in dir stream")
    proj.dir_raw = bytes(raw)
    proj.code_page = 949
    proj.dir_structure_dirty = True
    for m in proj.modules:
        m.dirty = True


def main() -> int:
    for name, path in MODULES:
        if not path.exists():
            print(f"ERROR: missing VBA module {path}", file=sys.stderr)
            return 1

    DIST.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        OUT.unlink()

    # 1) create a real macro-enabled workbook and add both modules
    xf = ExcelFile.create_new(str(OUT))
    proj = xf.vba_project()
    for name, path in MODULES:
        src = path.read_text(encoding="utf-8")
        proj.add_module(name, src)
    # remove the default empty Module1 created by create_new
    if "Module1" in proj.module_names():
        proj.delete_module("Module1")
    # auto-run on open: inject Workbook_Open into the ThisWorkbook document module
    _inject_auto_run(proj)
    # CRITICAL: force the VBA project code page to 949 (Korean) so 한글 in the
    # module source—including FUNCTIONAL literals like headerMap("\ucd9c\ubc1c\uc9c0ip")—
    # survives. pyOpenVBA's create_new template hardcodes PROJECTCODEPAGE=1252 in
    # dir_raw, and serialize_dir_stream reuses that prefix verbatim; modules are
    # then MBCS-encoded with project.code_page. Left at 1252, every 한글 becomes '?'
    # and the parser cannot match Korean headers at runtime in Excel.
    _force_codepage_949(proj)
    xf.save()
    xf.close()

    # 2) open with openpyxl preserving VBA, build/seed all sheets
    wb = openpyxl.load_workbook(str(OUT), keep_vba=True)

    # repurpose the default sheet as 'requests'
    base = wb[wb.sheetnames[0]]
    base.title = "requests"
    # Row 1: cosmetic group labels (출발지 over IP+설명, 목적지 over IP+설명).
    # Row 2: canonical leaf headers. Data from row 3. Mirrors VBA WriteRequestHeaders.
    col_index = {h: i + 1 for i, h in enumerate(REQUESTS_HEADERS)}
    src_ip_col = col_index["출발지"]
    src_desc_col = col_index["출발지설명"]
    dst_ip_col = col_index["목적지"]
    dst_desc_col = col_index["목적지설명"]
    base.cell(row=_REQ_HEADER_GROUP_ROW, column=src_ip_col, value="출발지")
    base.cell(row=_REQ_HEADER_GROUP_ROW, column=dst_ip_col, value="목적지")
    from openpyxl.utils import get_column_letter as _gcl
    base.merge_cells(f"{_gcl(src_ip_col)}{_REQ_HEADER_GROUP_ROW}:{_gcl(src_desc_col)}{_REQ_HEADER_GROUP_ROW}")
    base.merge_cells(f"{_gcl(dst_ip_col)}{_REQ_HEADER_GROUP_ROW}:{_gcl(dst_desc_col)}{_REQ_HEADER_GROUP_ROW}")
    for c, h in enumerate(REQUESTS_HEADERS, start=1):
        base.cell(row=_REQ_HEADER_ROW, column=c, value=h)
    tracking_rows: list[list[object]] = [list(REQUEST_TRACKING_HEADERS)]
    for r, example in enumerate(EXAMPLE_REQUEST_ROWS, start=_REQ_DATA_START_ROW):
        for header in REQUESTS_HEADERS:
            example_key = {"출발지": "출발지IP", "목적지": "목적지IP"}.get(header, header)
            val = example.get(example_key)
            if val is not None:
                base.cell(row=r, column=col_index[header], value=val)
        tracking_rows.append([
            r,
            example.get("원본파일", ""),
            example.get("원본행", ""),
            example.get("요청폴더", ""),
            example.get("제목", ""),
        ])
    # style the group label row too, then the leaf header row
    for c in (src_ip_col, dst_ip_col):
        gc = base.cell(row=_REQ_HEADER_GROUP_ROW, column=c)
        gc.font = _HEADER_FONT
        gc.fill = _HEADER_FILL
        gc.alignment = _HEADER_ALIGN
    _style_sheet(base, header_row=_REQ_HEADER_ROW)

    def add(title, rows):
        ws = wb.create_sheet(title)
        _write_rows(ws, rows)
        _style_sheet(ws)
        return ws

    add("firewalls", FIREWALLS)
    add("firewall_ranges", FIREWALL_RANGES)
    add("settings", SETTINGS)
    add("header_aliases", HEADER_ALIASES)
    add("processing_log", PROCESSING_LOG)
    add(REQUEST_TRACKING_SHEET, tracking_rows)
    add("route_results", [ROUTE_RESULTS_HEADERS])
    add("secui_cli", secui_cli_seed_rows())
    add("vendor_cli_templates", VENDOR_CLI_TEMPLATES)
    add("service_catalog", SERVICE_CATALOG)
    # sample-request-format has a blank A column header
    sf = wb.create_sheet("sample-request-format")
    _write_rows(sf, SAMPLE_FORMAT)
    _style_sheet(sf)
    add("usage", USAGE)
    add("주간보고", WEEKLY_REPORT)

    # build-time UX (input-assist / display only; after all sheets are seeded)
    apply_ux(wb)

    wb.save(str(OUT))
    wb.close()

    # 3) report
    size = os.path.getsize(OUT)
    print(f"Built {OUT} ({size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
