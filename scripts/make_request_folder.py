#!/usr/bin/env python3
"""Scaffold the request-folder/ tree the operator points the workbook at.

settings!request_folder is set to this folder (or a copy of it). Each team
subfolder is named <팀>_<문서번호>, which the merge macro parses into
request_team / request_doc_no. Drop request .xlsx files into the team folders,
then run the MergeFirewallRequestFolder macro.

Each .xlsx matches what the VBA parser reads:
  - first sheet, header row whose first labelled cell is `No` (col B; col A blank)
  - columns: 출발지IP 출발지 목적지IP 목적지 프로토콜 포트 방향 용도 시작일 종료일 비고
출발지IP/목적지IP accept a single IP, a CIDR block, or a ';'-separated list.

Run:
  ./.venv/bin/python scripts/make_request_folder.py
Output:
  request-folder/README.txt
  request-folder/_빈양식/신청서_빈양식.xlsx   (header-only template to copy)
  request-folder/<팀>_<문서번호>/*.xlsx
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "request-folder"

HEADER = [
    None, "No", "출발지IP", "출발지", "목적지IP", "목적지",
    "프로토콜", "포트", "방향", "용도", "시작일", "종료일", "비고",
]

# Real request entries. IPs sit inside the workbook's network_definitions zones
# (internal 10.10/16, dmz 10.20/16, server 172.16.1/24, outside 0.0.0.0/0) so the
# route engine resolves a concrete firewall path. Folder = <팀>_<문서번호>.
REQUESTS = [
    (
        "정보보호센터_1234",
        "신청서_웹서비스.xlsx",
        [
            [1, "10.10.10.5", "업무PC", "172.16.1.10", "업무시스템",
             "TCP", "443", "OUT", "HTTPS 업무 연동", "2026-01-01", "2026-12-31", "정기 신청"],
            [2, "10.10.20.0/24", "업무PC대역", "10.20.20.5", "DMZ서버",
             "TCP", "8080", "OUT", "내부->DMZ 연동", "2026-01-01", "2026-12-31", "대역 신청"],
        ],
    ),
    (
        "인프라팀_5678",
        "방화벽신청.xlsx",
        [
            [1, "10.10.30.10", "운영서버", "8.8.8.8", "외부DNS",
             "UDP", "53", "OUT", "외부 DNS 조회", "2026-02-01", "2026-12-31", ""],
        ],
    ),
    (
        "정보보호_2팀_9012",
        "신청.xlsx",
        [
            [1, "10.10.40.7", "관리PC", "172.16.1.20", "관리시스템",
             "TCP", "22", "OUT", "SSH 관리 접근", "2026-03-01", "2026-09-30", "한시 신청"],
            [2, "10.10.50.1;10.10.50.2", "관리PC군", "10.20.20.9", "DMZ관리",
             "TCP", "443", "OUT", "DMZ 관리 콘솔", "2026-03-01", "2026-09-30", "다중주소"],
        ],
    ),
]

# Header-only file the operator copies to start a new request workbook.
TEMPLATE_FOLDER = "_빈양식"
TEMPLATE_FILE = "신청서_빈양식.xlsx"

_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_HEADER_FONT = Font(bold=True, color="1F3864")
_CENTER = Alignment(horizontal="center", vertical="center")
_WIDTHS = [4, 6, 16, 12, 16, 12, 10, 8, 8, 18, 12, 12, 14]


def _write_request_file(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "firewall_requests"
    for c, val in enumerate(HEADER, start=1):
        if val is not None:
            ws.cell(row=1, column=c, value=val)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=2):  # data under `No` starts at col B
            ws.cell(row=r, column=c, value=val)
    for c in range(2, len(HEADER) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    for i, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


README = """\
request-folder — 신청서를 넣는 폴더

사용법
1) settings 시트의 request_folder 값에 이 폴더(또는 이 폴더를 복사한 경로)를 지정하거나
   매크로 SelectRequestFolder 로 이 폴더를 선택합니다.
2) 팀별 하위폴더(<팀명>_<문서번호>) 안에 신청서 .xlsx 를 넣습니다.
   - 폴더명 예: 정보보호센터_1234  ->  request_team=정보보호센터, request_doc_no=1234
   - 마지막 '_' 기준으로 팀/문서번호가 갈립니다(팀명에 '_'가 있어도 됨).
   - '_'가 없으면 폴더명 전체가 팀명, 문서번호는 빈값.
3) Excel 에서 매크로 MergeFirewallRequestFolder 실행(Alt+F8) -> 통합 + 라우팅 분석.

신청서 .xlsx 양식
- 첫 시트에 'No'(또는 '번호') 가 있는 행이 헤더 행입니다(이 폴더 파일은 B열에 No).
- 필요한 컬럼: 출발지IP 출발지 목적지IP 목적지 프로토콜 포트 방향 용도 시작일 종료일 비고
- 열 순서는 달라도 되고, 비표준 헤더는 settings!header_alias 로 매핑할 수 있습니다.
- 출발지IP/목적지IP 는 단일 IP, CIDR 대역, 세미콜론 구분 다중주소 모두 가능합니다.
- 새 신청서는 _빈양식/신청서_빈양식.xlsx 를 복사해서 작성하세요.

이 폴더에 들어 있는 신청 건은 워크북의 network_definitions zone 안의 IP를 사용하므로
그대로 통합/분석하면 실제 통과 방화벽 경로(target_firewalls)가 채워집니다.
"""


def main() -> int:
    if OUT.exists():
        shutil.rmtree(OUT)
    OUT.mkdir(parents=True)
    (OUT / "README.txt").write_text(README, encoding="utf-8")
    files = 0
    for folder, fname, rows in REQUESTS:
        _write_request_file(OUT / folder / fname, rows)
        files += 1
    _write_request_file(OUT / TEMPLATE_FOLDER / TEMPLATE_FILE, [])
    print(f"Built {OUT} ({files} request files + 1 빈 템플릿 in {len(REQUESTS)} team folders)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
